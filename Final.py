
import can
import time
import binascii
import sys
import os
import requests
import datetime
from datetime import datetime
import pytz
from PyQt5.QtGui  import QCursor
from PyQt5.QtCore import QPoint
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta,timezone
from PyQt5.QtWidgets import QMainWindow, QApplication,QMessageBox
from PyQt5.QtCore import QDateTime
from PyQt5.QtCore import QTimer
from finalTesting import Ui_FinalTestingUtility
import resources_rc

# Expected CAN IDs and their frame counts
expected_frame_counts = {0x100: 3, 0x101 :3, 0x103 : 4, 0x105 :3, 0x106 :3 , 0x115 : 1, 0x116 : 1,0x109 :1,0x110:1, 0x112 :2,0x113 :1,0x114:4
                         ,0x102 : 1,0x119 :1}

# Initialize received_frames with empty lists for each CAN ID
received_frames = {0x100: [],0x101 : [] , 0x103 :[],0x105 :[],0x106 :[] , 0x115 :[], 0x116 : [],0x109:[],0x110:[],0x112:[],0x113:[],0x114:[]
                   ,0x102:[] , 0x119 :[]}



class MyClass(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_FinalTestingUtility()
        self.ui.setupUi(self)
        self.stackedWidget = self.ui.stackedWidget
        self.bus = None
        self.busy = False
        #self.ui.pushButton.clicked.connect(self.goToPage2)
        self.ui.pushButton_8.clicked.connect(self.start_functions)
        self.ui.pushButton_2.clicked.connect(self.save_to_excel)
        self.ui.pushButton.clicked.connect(self.on_button_click)
        self.ui.pushButton_6.clicked.connect(self.DIs_func)
        self.initialize_can_bus()
        self.current_datetime =None
        self.IMEI_ascii= None
        self.ICCID_ascii = None
        self.appln_ver = None
        self.GSM_ver = None
        self.Gps_ver = None
        self.Int_vtg = None
        self.mains_vtg = None
        self.Gps_status = None
        self.No_of_Sat = None
        self.CREG = None
        self.CGREG = None
        self.CSQ = None
        self.operatorName =None
        self.MQTT_status = None
        self.No_of_LogInPacket =None
        self.frame1 = None
        self.frame2 = None
        self.frame3 = None
        self.frame4 = None
        self.frame5 =None
        self.converted_frame=None
        self.RTC = None
        self.EpochToCurrentTime =None
        self.Mains_result = None
        self.IntVtg_result = None
        self.Gps_result = None
        self.RTC_result =None
        self.GSM_result = None
        self.DIs_result = None
        self.IGN_result = None
        self.Tamper_result = None
        self.MEMS_result = None
        self.MQTT_result = None
        self.DI1_H =None
        self.DI1_L= None
        self.IGN = None
        self.DI1_status = False
        self.DI2_status = False
        self.DI3_status = False
        self.DI1_seen_0 = False
        self.DI1_seen_1 = False
        self.DI2_seen_0 = False
        self.DI2_seen_1 = False
        self.DI3_seen_0 = False
        self.DI3_seen_1 = False
        self.CREG_found = False
        self.CGREG_found = False
        self.CSQ_found = False
        self.operator_found = False
        self.IGN_seen_0 = False
        self.IGN_seen_1 = False
        self.No_of_Sat2 = None
        self.concatenated_hex = None
        self.concatenated_satellites_decimal = None



        self.stage4_url = "http://192.168.2.253:6101/api/stage4"
        self.stage5_url = "http://192.168.2.253:6101/api/stage5"
        self.device_status_url = "http://192.168.2.253:6101/api/test_points/SENSOR001"
        self.check_server_url = "http://192.168.2.253:6101/api/PROD_check"
        
        self.barcode = "SENSOR001"
        self.model_name = None
        # Initialize a QTimer
        self.timer = QTimer(self)
        self.timer.setInterval(1000)  # 1000 ms = 1 second
        self.timer.timeout.connect(self.on_timer_timeout)
        self.elapsed_time = 0
        self.operator = None
        self.qc_head = None
        # Initialize CAN bus in the __init__ method to avoid reinitializing it every time
        cursor = QCursor()
        cursor.setPos(620, 70)

        self.data = {
            "QR_code": self.barcode,
            "model_name": "ACON4L",
            "final_testing_status": True,
            "IMEI": '123456789',
            "ICCID": "89911234567890123456",
            "SystemRtc": "2024-03-21 15:30:45",
            "AppFWVersion": "1.0.5",
            "BLFWVersion": "2.1.0",
            "GPSFWVersion": "3.2.1",
            "GSMFWVersion": "4.0.2",
            "HWVersion": "V2.0",
            "GPSFix": "3D Fix",
            "HDOP": "1.2",
            "PDOP": "2.4",
            "No_satelite": "8",
            "GSMStatus": "Connected",
            "signalStrength": "-67",
            "Network_code": "40414",
            "Network_Type": "4G",
            "SIM": "Active",
            "MEMS": "Working",
            "Voltage": "12.5",
            "Memory": "75",
            "Ignition": "ON",
            "Tamper": "No",
            "DI_1_H": "1",
            "DI_1_L": "0",
            "DI_2_H": "1",
            "DI_2_L": "0",
            "DI_3_H": "1",
            "DI_3_L": "0",
            "DO_1_H": "1",
            "DO_1_L": "0",
            "DO_2_H": "1",
            "DO_2_L": "0",
            "CAN": "OK",
            "RS485": "Connected",
            "AnalogInput1": "4.5",
            "AnalogInput2": "3.2",
        }
        self.headers = {"Content-Type": "application/json"}

        # Initialize flags
        self.function100_done = False
        self.function101_done = False
        self.function103_done = False
        self.function105_done = False
        self.function106_done = False
        self.function115_done = False
        self.function116_done = False
        self.function109_done = False
        self.function110_done = False
        self.function112_done = False
        self.function113_done = False
        self.function114_done = False
        self.function102_done = False

        # Timer for delays
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)  # Ensure it only fires once
        self.timer.timeout.connect(self.execute_next_function)

        

    def initialize_can_bus(self):
        try:
            # Initialize the bus once, not inside each function
            self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
            #print(f"CAN Bus initialized: {self.bus.channel_info}")
        except can.CanError as e:
            print(f"Error initializing CAN bus: {str(e)}")
            self.bus = None  # Set bus to None if there's an initialization error

    def start_functions(self):
        """Start the series of functions when the button is clicked."""
        #print("Button clicked. Starting functions.")
        
        # Reset flags
        self.function100_done = False
        self.function101_done = False
        self.function103_done = False
        self.function105_done = False
        self.function106_done = False
        self.function115_done = False
        self.function116_done = False
        self.function109_done = False
        self.function110_done = False
        self.function112_done = False
        self.function113_done = False
        self.function114_done = False
        self.function102_done = False
        
        # Call the first function
        self.fun_0x103()

    def fun_0x100(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x100, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x100]):
                message = self.bus.recv(timeout=2)  # 2 second timeout for each frame
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x100].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x100]) == expected_frame_counts[0x100]:
                frames = received_frames[0x100]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Reassembled message for CAN ID 0x100: {complete_message.hex()}")

                IMEI = complete_message[:15]
                #print(f"Extracted IMEI: {IMEI.hex()}")
 
                try:
                  self.IMEI_ascii = IMEI.decode('ascii')  # Decode bytes into ASCII string
                  
                  print(f"Extracted IMEI (ASCII): {self.IMEI_ascii}")
                  self.ui.plainTextEdit_10.setPlainText(self.IMEI_ascii)
                  self.ui.plainTextEdit_12.appendPlainText(f"IMEI : {self.IMEI_ascii}\n")

                  if len(self.IMEI_ascii) < 15:
                     self.ui.plainTextEdit_10.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_10.setStyleSheet("background-color: white;")
                  
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x100. Expected {expected_frame_counts[0x100]}, but received {len(received_frames[0x100])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x100].clear()
            self.function100_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x100")

    def fun_0x101(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x101, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x101]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                print("hex ICCID received :",message)
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x101].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x101]) == expected_frame_counts[0x101]:
                frames = received_frames[0x101]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Reassembled message for CAN ID 0x101: {complete_message.hex()}")
 
                ICCID = complete_message[:20]
                #print(f"Extracted IMEI: {ICCID.hex()}")
 
                try:
                  self.ICCID_ascii = ICCID.decode('ascii')  # Decode bytes into ASCII string
                  print(f"Extracted ICCID (ASCII): {self.ICCID_ascii}")
                  self.ui.plainTextEdit_11.setPlainText(self.ICCID_ascii)
                  self.ui.plainTextEdit_12.appendPlainText(f"ICCID : {self.ICCID_ascii}\n")

                  if len(self.ICCID_ascii)<20:
                      self.ui.plainTextEdit_11.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_11.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x101. Expected {expected_frame_counts[0x101]}, but received {len(received_frames[0x101])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x101].clear()
            self.function101_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x101")

    def fun_0x103(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x103, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x103]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x103].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x103. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x103]) == expected_frame_counts[0x103]:
                frames = received_frames[0x103]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Reassembled message for CAN ID 0x100: {complete_message.hex()}")
 
                try:
                  self.appln_ver = complete_message.decode('ascii').rstrip('\x00').strip() # Decode bytes into ASCII string
                  #print('appln ver ASCII :',self.appln_ver)
                  print(f"Application version: {repr(self.appln_ver)}")
                  self.ui.plainTextEdit_8.setPlainText(self.appln_ver)
                  self.ui.plainTextEdit_12.appendPlainText(f"Application Version : {self.appln_ver}\n")

                  if self.appln_ver != 'SAM01_LITE_PROD_0.0.1_TST01':
                      self.ui.plainTextEdit_8.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_8.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x103. Expected {expected_frame_counts[0x103]}, but received {len(received_frames[0x103])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x103].clear()
            self.function103_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x103")

    def fun_0x105(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x105, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x105]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x105].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x105]) == expected_frame_counts[0x105]:
                frames = received_frames[0x105]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Complete for CAN ID 0x105: {complete_message.hex()}")
 
                try:
                  self.Gps_ver = complete_message.decode('ascii').rstrip('\x00').strip()  # Decode bytes into ASCII string
                  print('GPS ver ASCII :',self.Gps_ver)
                  self.ui.plainTextEdit_5.setPlainText(self.Gps_ver)
                  self.ui.plainTextEdit_12.appendPlainText(f"GPS Version : {self.Gps_ver}\n")
                  #gps_ver_cleaned = self.Gps_ver.strip()
                  print('gps strip',{repr(self.Gps_ver)})
                  
                  if self.Gps_ver != 'EC200UCNAAR03A11M08':
                      self.ui.plainTextEdit_5.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_5.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x105. Expected {expected_frame_counts[0x105]}, but received {len(received_frames[0x105])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x105].clear()
            self.function105_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x105")

    def fun_0x106(self):
        print('fun 0x106 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x106, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x106]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x106].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x106]) == expected_frame_counts[0x106]:
                frames = received_frames[0x106]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Complete message for CAN ID 0x106: {complete_message.hex()}")
 
                try:
                  self.GSM_ver = complete_message.decode('ascii').rstrip('\x00').strip()
                  print('GSM ver ASCII :',{repr(self.GSM_ver)})
                  self.ui.plainTextEdit_6.setPlainText(self.GSM_ver)
                  self.ui.plainTextEdit_12.appendPlainText(f"GSM Version : {self.GSM_ver}\n")

                  if self.GSM_ver != 'EC200UCNAAR03A11M08':
                      self.ui.plainTextEdit_6.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_6.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x106. Expected {expected_frame_counts[0x106]}, but received {len(received_frames[0x106])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x106].clear()
            self.function106_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x106")

    def fun_0x115(self):
        print('115 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x115, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response
            print('hex value of mains :', message)

            if message:
                exttracted_mains = message.data[1:]
                # Process the received message
                #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
            
                # Decode the received message and update the UI
                self.mains_vtg = exttracted_mains.decode('ascii')  # Decode bytes into ASCII string
                print('Mains vtg:', self.mains_vtg)
            
                # Update the UI with the decoded message
                self.ui.mains_input_2.setPlainText(self.mains_vtg)
                self.ui.plainTextEdit_12.appendPlainText(f"Mains Voltage: {self.mains_vtg}\n")

                if self.mains_vtg > '24.00':
                      self.ui.mains_input_2.setStyleSheet("background-color: red;")
                      self.ui.plainTextEdit_29.setPlainText("Fail")
                      self.ui.plainTextEdit_29.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:red""")
                else:
                      self.ui.mains_input_2.setStyleSheet("background-color: white;")
                      self.ui.plainTextEdit_29.setPlainText("Pass")
                      self.ui.plainTextEdit_29.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")

                self.Mains_result= self.ui.plainTextEdit_29.toPlainText()
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x115. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x115].clear()
            self.function115_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x115")

    def fun_0x116(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x116, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                exttracted_IntVtg = message.data[1:]
                # Process the received message
                #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
            
                # Decode the received message and update the UI
                self.Int_vtg = exttracted_IntVtg.decode('ascii')  # Decode bytes into ASCII string
                print('Internal vtg:', self.Int_vtg)
            
                # Update the UI with the decoded message
                self.ui.IntBat_input_2.setPlainText(self.Int_vtg)
                self.ui.plainTextEdit_12.appendPlainText(f"Int_Bat Voltage: {self.Int_vtg}\n")

                if self.Int_vtg > '4.2':
                      self.ui.IntBat_input_2.setStyleSheet("background-color: red;")
                      self.ui.plainTextEdit_23.setPlainText("Fail")
                      self.ui.plainTextEdit_23.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:red""")
                else:
                      self.ui.IntBat_input_2.setStyleSheet("background-color: white;")
                      self.ui.plainTextEdit_23.setPlainText("Pass")
                      self.ui.plainTextEdit_23.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")

                self.IntVtg_result = self.ui.plainTextEdit_23.toPlainText()
                
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x116. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x116].clear()
            self.function116_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x116")

    def fun_0x109(self):
        print('109 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x109, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                # Get GPS status and display it
                self.Gps_status = message.data[1]
                print('Gps status :', self.Gps_status)
                self.ui.Operator_2.setPlainText(str(self.Gps_status))
    
                # Get the two hex values for the number of satellites
                self.No_of_Sat = hex(message.data[2])
                print('1st byte of no. of sat', self.No_of_Sat)
                self.No_of_Sat2 = hex(message.data[3])
                print('2nd byte of no. of sat', self.No_of_Sat2)
    
                # Concatenate the two hex values (removing the '0x' part before concatenation)
                self.concatenated_hex = self.No_of_Sat[2:] + self.No_of_Sat2[2:]
                print('Concatenated hex value:', self.concatenated_hex)
    
                # Convert the concatenated hex string to a decimal value
                self.concatenated_satellites_decimal = int(self.concatenated_hex, 16)
                print('No. of sat (Decimal):', self.concatenated_satellites_decimal)
    
                # Update the UI with GPS status and number of satellites
                self.ui.plainTextEdit_12.appendPlainText(f"GPS status: {str(self.Gps_status)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"No. of Satellite: {str(self.concatenated_satellites_decimal)}\n")
    
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

            if self.Gps_status != 1:
                self.ui.Operator_2.setStyleSheet("background-color:red")
                self.ui.plainTextEdit_28.setPlainText("Fail")
                self.ui.plainTextEdit_28.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:red""")
            else:
                
                self.ui.plainTextEdit_28.setPlainText("Pass")
                self.ui.plainTextEdit_28.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")

            if self.concatenated_satellites_decimal ==3000 or self.concatenated_satellites_decimal == 184:
                self.ui.NoOf_satellite.setPlainText(str('0'))
                self.ui.NoOf_satellite.setStyleSheet("background-color: red;")
            else:
                self.ui.NoOf_satellite.setPlainText(str(self.concatenated_satellites_decimal))
                self.ui.NoOf_satellite.setStyleSheet("background-color: white;")

            self.Gps_result = self.ui.plainTextEdit_28.toPlainText()




        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x109].clear()
            self.function109_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x109")

    

    def fun_0x110(self):
        print('inside 110')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x110, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                self.CREG = message.data[1]
                print('CREG :',self.CREG)
                self.ui.CSQ.setPlainText(str(self.CREG))
                if self.CREG !=5:
                    self.ui.CSQ.setStyleSheet("background-color: red;")
                else:
                    self.ui.CSQ.setStyleSheet("background-color: white;")
                    self.CREG_found = True

                self.CGREG = message.data[2]
                print('CGREG:',self.CGREG)
                self.ui.CGREG.setPlainText(str(self.CGREG))
                if self.CGREG != 5:
                    self.ui.CGREG.setStyleSheet("background-color: red;")
                else:
                    self.ui.CGREG.setStyleSheet("background-color: white;")
                    self.CGREG_found =True

                self.CSQ =message.data[3]
                print('CSQ',self.CSQ)
                self.ui.CREG.setPlainText(str(self.CSQ))
                if self.CSQ < 10:
                    self.ui.CREG.setStyleSheet("background-color: red;")
                else:
                    self.ui.CREG.setStyleSheet("background-color: white;")
                    self.CSQ_found =True

                
                print(self.CREG_found, self.CGREG_found, self.CSQ_found)
                if self.CREG_found and self.CGREG_found:
                    self.ui.plainTextEdit_24.setPlainText("Pass")
                    self.ui.plainTextEdit_24.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
                else:
                    self.ui.plainTextEdit_24.setPlainText("Fail")
                    self.ui.plainTextEdit_24.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

                self.ui.plainTextEdit_12.appendPlainText(f"CREG: {str(self.CREG)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"CGREG: {str(self.CGREG)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"CSQ : {str(self.CSQ)}\n")

                self.GSM_result = self.ui.plainTextEdit_24.toPlainText()
            
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

           

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x110].clear()
            self.function110_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x109")

    def fun_0x112(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x112, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x112]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x112].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x112]) == expected_frame_counts[0x112]:
                frames = received_frames[0x112]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Complete for CAN ID 0x112: {complete_message.hex()}")
 
                try:
                  self.operatorName = complete_message.decode('ascii').rstrip('\x00').strip()  # Decode bytes into ASCII string
                  print('Operator Name :',{repr(self.operatorName)})
                  self.ui.Operator.setPlainText(self.operatorName)
                  self.ui.plainTextEdit_12.appendPlainText(f"Operator Name : {self.operatorName}\n")
                 
                  
                  if self.operatorName == 'AIRTEL 4G':
                      self.ui.Operator.setStyleSheet("background-color: white;")
                      self.operator_found = True
                  else:
                      self.ui.Operator.setStyleSheet("background-color: red;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x112. Expected {expected_frame_counts[0x112]}, but received {len(received_frames[0x112])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x112].clear()
            self.function112_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x105")

    def fun_0x113(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x113, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                self.MQTT_status = message.data[1]
                print('MQTT status :',self.MQTT_status)

                if self.MQTT_status == 0 or self.MQTT_status == 255:
                    self.ui.Analog1_2.setStyleSheet("background-color: red;")
                    self.ui.Analog1_2.setPlainText(str('0'))
                    self.ui.plainTextEdit_31.setPlainText("Fail")
                    self.ui.plainTextEdit_31.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
                else:
                    self.ui.Analog1_2.setStyleSheet("background-color: white;")
                    self.ui.Analog1_2.setPlainText(str(self.MQTT_status))
                    self.ui.plainTextEdit_31.setPlainText("Pass")
                    self.ui.plainTextEdit_31.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

                self.No_of_LogInPacket = message.data[2]
                print('No. of Login Packet:',self.No_of_LogInPacket)
                self.ui.MQTT.setPlainText(str(self.No_of_LogInPacket))
                if self.No_of_LogInPacket != 1:
                    self.ui.MQTT.setStyleSheet("background-color: red;")
                else:
                    self.ui.MQTT.setStyleSheet("background-color: white;")

                self.ui.plainTextEdit_12.appendPlainText(f"MQTT status: {str(self.MQTT_status)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"No. of Login packet: {str(self.No_of_LogInPacket)}\n")
            
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

            self.MQTT_result = self.ui.plainTextEdit_31.toPlainText()

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x112].clear()
            self.function113_done = True
            time.sleep(2)
            self.execute_next_function()
            #print("Frames cleared for CAN ID 0x109")

    def fun_0x114(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy
        try:
            # Create and send the message
            msg = can.Message(arbitration_id=0x114, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
            self.bus.send(msg)

            # Reset frames before starting to receive
            self.frame1 = None
            self.frame2 = None
            self.frame3 = None

            # Create a list to keep track of frames received so far
            received_frames = 0

            # Wait for the response
            all_frames = []  # Initialize an empty list to collect all frames

            # Try receiving all frames
            for i in range(expected_frame_counts[0x114]):  # Ensure it defaults to 0 if the key is missing
                message = self.bus.recv(timeout=2)  # 2-second timeout for each frame
                print('recived frames',message)
                if message:
                    # Extract the frame, skipping the 0th byte
                    frame = message.data[1:].hex()  # Skip the 0th byte and convert the rest to hex
                    all_frames.append(frame)  # Append the frame to the list

            # Join all frames into a single string and print
            frames = ''.join(all_frames)
            print("All Frames:", frames)

            # Convert the concatenated hex frames into bytes
            byte_data = bytes.fromhex(frames)

            # Decode the byte data into an ASCII string
            ascii_string = byte_data.decode('ascii', errors='ignore')
            print("ascii_string:", ascii_string)

            # Store the ASCII string in 3 frames, each having a part of the string
            frame_size = len(ascii_string) // 3  # Divide the string into 3 equal parts

            # Split the string into three parts, ensuring no part is empty
            self.frame1 = ascii_string[:frame_size]
            self.frame2 = ascii_string[frame_size:frame_size*2]
            self.frame3 = ascii_string[frame_size*2:]  # The remaining part goes into frame3

            # Print the frames for verification
            print("Frame 1:", self.frame1)
            print("Frame 2:", self.frame2)
            print("Frame 3:", self.frame3)

            self.ui.MEMS_Xa.setPlainText(self.frame1)
            self.ui.MEMS_Ya.setPlainText(self.frame2)
            self.ui.MEMS_Za.setPlainText(self.frame3)
            self.ui.plainTextEdit_12.appendPlainText(f"Accelerometer data: {self.frame1}, {self.frame2}, {self.frame3}")

            if self.frame1 and self.frame2 and self.frame3:
                self.ui.plainTextEdit_25.setPlainText("Pass")
                self.ui.plainTextEdit_25.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                self.ui.plainTextEdit_25.setPlainText("Fail")
                self.ui.plainTextEdit_25.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.MEMS_result = self.ui.plainTextEdit_25.toPlainText()

        except Exception as e:
            # Handle any exceptions that occur during the process
            print(f"An error occurred: {e}")

        # Optionally, set `self.busy` to False if an error occurs to allow retries
            self.busy = False

            
        finally:
            self.busy = False  # Mark the system as not busy
            self.function114_done = True
            time.sleep(2)
            self.execute_next_function()





    def fun_0x102(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy
        current_time_utc = datetime.now(pytz.utc)  # Get current time in UTC (offset-aware)
        current_time_ist = current_time_utc.astimezone(pytz.timezone('Asia/Kolkata'))  # Convert to IST
        print('Current time (IST):', current_time_ist)

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x102, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                RTC_data = message.data[1:5]
                #print('Hex RTC:', RTC_data)

                self.RTC = int.from_bytes(RTC_data, byteorder='big')
                #print('INT RTC:', self.RTC)

                # Convert the epoch time (RTC) to datetime in UTC
                epoch_time_utc = datetime.fromtimestamp(self.RTC, tz=pytz.utc)
                epoch_time_ist = epoch_time_utc.astimezone(pytz.timezone('Asia/Kolkata'))  # Convert to IST
                print('Epoch to current time (IST):', epoch_time_ist)

                # Compare the two times (both are now offset-aware)
                time_difference = current_time_ist - epoch_time_ist


                # Compare the two times (both are now offset-aware)
                if time_difference.total_seconds()>15:
                    self.ui.plainTextEdit_21.setPlainText("Fail")
                    self.ui.plainTextEdit_21.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
                else:
                    self.ui.plainTextEdit_21.setPlainText("Pass")
                    self.ui.plainTextEdit_21.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x102. No response received.")

            self.RTC_result = self.ui.plainTextEdit_21.toPlainText()

        except can.CanError as e:
            print(f"CAN error: {str(e)}")

        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x102].clear()  # Clear any frames in the buffer for ID 0x102
            self.function102_done = True
            time.sleep(2)  # Sleep to allow processing
            self.execute_next_function()  # Move on to the next function

    def DIs_func(self):
            if self.busy:  # Check if the system is busy
                print("System is busy, please wait...")
                return

            if self.bus is None:  # Check if the bus was initialized properly
                print("CAN Bus not initialized. Cannot send message.")
                return

            self.busy = True  # Mark the system as busy

            try:
                # Create the CAN message
                msg = can.Message(arbitration_id=0x119, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

                # Send the message once
                self.bus.send(msg)

                # Wait for a response with a timeout (e.g., 2 seconds)
                message = self.bus.recv(timeout=2)  # 2 seconds timeout for response
        
                if message:
                    self.IGN = message.data[1]
                    print('IGN :', self.IGN)
            
                    self.tamper = message.data[2]
                    self.ui.Tamp_L.setPlainText(str(self.tamper))
                    print('Tamper:', self.tamper)
            
                    self.DI1 = message.data[3]
                    print('DI1 :', self.DI1)
            
                    self.DI2 = message.data[4]
                    print('DI2 :', self.DI2)

                    self.DI3 = message.data[5]
                    print('DI3 :', self.DI3)
            
                else:
                    # If no message is received within the timeout period
                    print(f"Timeout waiting for message for CAN ID 0x119. No response received.")

                if self.tamper != 0:
                    self.ui.Tamp_L.setStyleSheet("background-color : red")
                    self.ui.plainTextEdit_26.setPlainText("Fail")
                    self.ui.plainTextEdit_26.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
                else:
                    self.ui.Tamp_L.setStyleSheet("background-color : white")
                    self.ui.plainTextEdit_26.setPlainText("Pass")
                    self.ui.plainTextEdit_26.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

                self.Tamper_result = self.ui.plainTextEdit_26.toPlainText()

                # Update UI fields if they are empty
                if not self.ui.DI1_H_3.toPlainText():
                    self.ui.DI1_H_3.setPlainText(str(self.IGN))
                else:
                    self.ui.IGN_H.setPlainText(str(self.IGN))

                if not self.ui.DI1_H_6.toPlainText():
                    self.ui.DI1_H_6.setPlainText(str(self.DI1))
                else:
                    self.ui.DI1_H_7.setPlainText(str(self.DI1))

                if not self.ui.DI1_H_4.toPlainText():
                    self.ui.DI1_H_4.setPlainText(str(self.DI2))
                else:
                    self.ui.DI1_H_5.setPlainText(str(self.DI2))

                if not self.ui.DI1_H_8.toPlainText():
                    self.ui.DI1_H_8.setPlainText(str(self.DI3))
                else:
                    self.ui.DI_H.setPlainText(str(self.DI3))

                self.ui.plainTextEdit_12.appendPlainText(f"IGN: {str(self.IGN)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"Tamper: {str(self.tamper)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"DI1,DI2,DI3: {self.DI1}, {self.DI2}, {self.DI3}")

                
                # Track whether we have seen both states (0 and 1) for each DI
                # Check and update DI1 status
                if self.DI1 == 0:
                    self.DI1_seen_0 = True  # Mark that we have seen 0 for DI1
                elif self.DI1 == 1:
                    self.DI1_seen_1 = True  # Mark that we have seen 1 for DI1

                # Check and update DI2 status
                if self.DI2 == 0:
                    self.DI2_seen_0 = True  # Mark that we have seen 0 for DI2
                elif self.DI2 == 1:
                    self.DI2_seen_1 = True  # Mark that we have seen 1 for DI2

                # Check and update DI3 status
                if self.DI3 == 0:
                    self.DI3_seen_0 = True  # Mark that we have seen 0 for DI3
                elif self.DI3 == 1:
                    self.DI3_seen_1 = True  # Mark that we have seen 1 for DI3

                if self.IGN == 0:
                    self.IGN_seen_0 =True
                else:
                    self.IGN_seen_1 = True

                # Use QTimer to periodically check if both 0 and 1 have been seen for each DI
                self.timer = QTimer(self)
                self.timer.timeout.connect(self.check_flags)  # Connect timeout to the check_flags function
                self.timer.start(1000)  # Check every second (1000 ms)

            except can.CanError as e:
                print(f"CAN error: {str(e)}")
    
            finally:
                self.busy = False  # Mark the system as not busy
                received_frames[0x119].clear()
                time.sleep(2)

    def check_flags(self):
        # This method will be called every second
        #print(f"Checking flags: DI1_seen_0={self.DI1_seen_0}, DI1_seen_1={self.DI1_seen_1}, DI2_seen_0={self.DI2_seen_0}, DI2_seen_1={self.DI2_seen_1}, DI3_seen_0={self.DI3_seen_0}, DI3_seen_1={self.DI3_seen_1}")

        # Now check if all flags are True
        if self.DI1_status and self.DI2_status and self.DI3_status:
            self.timer.stop()  # Stop the timer when all flags are True
            print('timer stopped')
        
            # Now that all DI states are confirmed (both 0 and 1), determine the result
            if self.DI1_status and self.DI2_status and self.DI3_status:
                self.ui.plainTextEdit_22.setPlainText("Pass")
                self.ui.plainTextEdit_22.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                self.ui.plainTextEdit_22.setPlainText("Fail")
                self.ui.plainTextEdit_22.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
        

        if self.IGN_seen_0 and self.IGN_seen_1:
            self.timer.stop()
            if self.IGN_seen_0 and self.IGN_seen_1:

                self.ui.plainTextEdit_30.setPlainText("Pass")
                self.ui.plainTextEdit_30.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                self.ui.plainTextEdit_30.setPlainText("Fail")
                self.ui.plainTextEdit_30.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
    
        self.DIs_result = self.ui.plainTextEdit_22.toPlainText()
        self.IGN_result = self.ui.plainTextEdit_30.toPlainText()

        
    


    def execute_next_function(self):
        """Check which function is done and call the next one."""
        if self.function103_done and not self.function106_done:
            self.fun_0x106()  # Call function 2 after function 1 is done

        elif self.function106_done and not self.function105_done:
             self.fun_0x105()  # Call function 3 after function 2 is done

        elif self.function105_done and not self.function101_done:
            self.fun_0x101()

        elif self.function101_done and not self.function100_done:
            self.fun_0x100()

        elif self.function100_done and not self.function110_done:
            self.fun_0x110()

        elif self.function110_done and not self.function112_done:
            self.fun_0x112()

        elif self.function112_done and not self.function109_done:
            self.fun_0x109()

        elif self.function109_done and not self.function115_done:
            self.fun_0x115()

        elif self.function115_done and not self.function116_done:
            self.fun_0x116()

        elif self.function116_done and not self.function113_done:
            self.fun_0x113()

        elif self.function113_done and not self.function114_done:
            self.fun_0x114()

        elif self.function114_done and not self.function102_done:
            self.fun_0x102()
        else:
            print("All functions completed.")
            # You can enable a button or perform other tasks once all functions are done
            self.ui.pushButton_2.setEnabled(True)  # Enable button after all functions are done

    


 
 
    def login(self):
        # Get the username and password entered by the user
        Operator = self.ui.plainTextEdit.toPlainText()
        QC_head = self.ui.plainTextEdit_2.toPlainText()
 
        # Logic to check the username and password
        if Operator is not None and QC_head is not None:
            self.ui.pushButton.clicked.connect(self.goToPage2)
            self.show_message("Login Successful", "Welcome, Operator!")
        else:
            self.show_message("Login Failed", "Invalid username or password. Please try again.")
 
    def goToPage2(self):
        self.stackedWidget.setCurrentIndex((self.stackedWidget.currentIndex()+1)%2)

    def clean_string(self,input_string):
    
        # Filter out characters with ASCII values from 0-31 and 127
        return ''.join(c for c in input_string if 31 < ord(c) < 127)

    def save_to_excel(self):
        # Get the current date in YYYY-MM-DD format
        current_date = datetime.now().strftime("%Y-%m-%d")

        # Create a file path with the date appended to the filename
        current_directory = os.getcwd()
        file_path = os.path.join(current_directory, f"Final_Testing_DeviceStatus_{current_date}.xlsx")

        # If the file exists, load the existing workbook; otherwise, create a new one
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

            # Set the headers for the columns (only if the file is being created for the first time)
            headers = ['Date', 'Model', 'Operator', 'QC Head', 'IMEI', 'ICCID', 'Application Version', 'GSM Version',
                   'GPS Version', 'Mains vtg', 'Int_Bat vtg', 'GPS status', 'No.of Sat', 'CREG', 'CGREG', 'CSQ',
                   'Operator', 'MQTT', 'No.Of Login packet', 'MEMS Xa', 'MEMS Ya', 'MEMS Za', 'Mains result', 'IntBat result'
                   ,'Gps result', 'RTC','GSM result','DIs result','IGN result','Tamper result','MEMS result','MQTT result']
            ws.append(headers)  # Append headers as the first row

    # Clean the data before inserting into the worksheet
        data = [
        self.clean_string(str(self.current_datetime)) if self.current_datetime is not None else 'Not found',
        self.clean_string(str(self.model_name)) if self.model_name is not None else 'Not found',
        self.clean_string(self.operator) if self.operator is not None else 'Not found',
        self.clean_string(self.qc_head) if self.qc_head is not None else 'Not found',
        self.clean_string(self.IMEI_ascii) if self.IMEI_ascii is not None else 'Not found',
        self.clean_string(self.ICCID_ascii) if self.ICCID_ascii is not None else 'Not found',
        self.clean_string(self.appln_ver) if self.appln_ver is not None else 'Not found',
        self.clean_string(self.GSM_ver) if self.GSM_ver is not None else 'Not found ',
        self.clean_string(self.Gps_ver) if self.Gps_ver is not None else 'Not found',
        self.clean_string(self.mains_vtg) if self.mains_vtg is not None else 'Not found',
        self.clean_string(self.Int_vtg) if self.Int_vtg is not None else 'Not found',
        self.clean_string(str(self.Gps_status)) if self.Gps_status is not None else 'Not found',
        self.clean_string(str(self.No_of_Sat)) if self.No_of_Sat is not None else 'Not found',
        self.clean_string(str(self.CREG)) if self.CREG is not None else 'Not found',
        self.clean_string(str(self.CGREG)) if self.CGREG is not None else 'Not found',
        self.clean_string(str(self.CSQ)) if self.CSQ is not None else 'Not found',
        self.clean_string(str(self.operatorName)) if self.operatorName is not None else 'Not found',
        self.clean_string(str(self.MQTT_status)) if self.MQTT_status is not None else 'Not found',
        self.clean_string(str(self.No_of_LogInPacket)) if self.No_of_LogInPacket is not None else 'Not found',
        self.clean_string(str(self.frame1)) if self.frame1 is not None else 'Not found',
        self.clean_string(str(self.frame2)) if self.frame2 is not None else 'Not found',
        self.clean_string(str(self.frame5)) if self.frame5 is not None else 'Not found',
        self.clean_string(str(self.Mains_result)) if self.Mains_result is not None else 'Not found',
        self.clean_string(str(self.IntVtg_result)) if self.IntVtg_result is not None else 'Not found',
        self.clean_string(str(self.Gps_result)) if self.Gps_result is not None else 'Not found',
        self.clean_string(str(self.RTC_result)) if self.RTC_result is not None else 'Not found',
        self.clean_string(str(self.GSM_result)) if self.GSM_result is not None else 'Not found',
        self.clean_string(str(self.DIs_result)) if self.DIs_result is not None else 'Not found',
        self.clean_string(str(self.IGN_result)) if self.IGN_result is not None else 'Not found',
        self.clean_string(str(self.Tamper_result)) if self.Tamper_result is not None else 'Not found',
        self.clean_string(str(self.MEMS_result)) if self.MEMS_result is not None else 'Not found',
        self.clean_string(str(self.MQTT_result)) if self.MQTT_result is not None else 'Not found'
        ]

        # Append the data to the next available row
        ws.append(data)

        try:
            # Save the workbook to the specified path
            wb.save(file_path)
            print(f"Data saved to {file_path}")
            QMessageBox.information(self, "Success", f"Data successfully saved to {file_path}")

            # Clear the UI inputs after saving data
            self.clear_ui()

        except Exception as e:
            print(f"Error saving data to Excel: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to save data to Excel: {str(e)}")


    def clear_ui(self):
        
        self.ui.QC_input_2.clear()
        self.ui.QC_input_2.setStyleSheet("background-color: white;")
        self.ui.barcode_Input_2.clear()
        self.ui.barcode_Input.setStyleSheet("background-color: white;")
        self.ui.server_Input.clear()
        self.ui.server_Input.setStyleSheet("background-color: white;")
        self.ui.barcode_Input.clear()
        self.ui.barcode_Input.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_5.clear()
        self.ui.plainTextEdit_5.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_6.clear()
        self.ui.plainTextEdit_6.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_8.clear()
        self.ui.plainTextEdit_8.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_9.clear()
        self.ui.plainTextEdit_9.setStyleSheet("background-color: white;")
        self.ui.Tamp_L.clear()
        self.ui.Tamp_L.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_10.clear()
        self.ui.plainTextEdit_10.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_11.clear()
        self.ui.plainTextEdit_11.setStyleSheet("background-color: white;")
        self.ui.CSQ.clear()
        self.ui.CSQ.setStyleSheet("background-color: white;")
        self.ui.CGREG.clear()
        self.ui.CGREG.setStyleSheet("background-color: white;")
        self.ui.CREG.clear()
        self.ui.CREG.setStyleSheet("background-color: white;")
        self.ui.Operator.clear()
        self.ui.Operator.setStyleSheet("background-color: white;")
        self.ui.NoOf_satellite.clear()
        self.ui.NoOf_satellite.setStyleSheet("background-color: white;")
        self.ui.Operator_2.clear()
        self.ui.Operator_2.setStyleSheet("background-color: white;")
        self.ui.MQTT.clear()
        self.ui.MQTT.setStyleSheet("background-color: white;")
        self.ui.mains_input_2.clear()
        self.ui.mains_input_2.setStyleSheet("background-color: white;")
        self.ui.IntBat_input_2.clear()
        self.ui.IntBat_input_2.setStyleSheet("background-color: white;")
        self.ui.Analog1_2.clear()
        self.ui.Analog1_2.setStyleSheet("background-color: white;")
        self.ui.MEMS_Xa.clear()
        self.ui.MEMS_Xa.setStyleSheet("background-color: white;")
        self.ui.MEMS_Ya.clear()
        self.ui.MEMS_Ya.setStyleSheet("background-color: white;")
        self.ui.MEMS_Za.clear()
        self.ui.MEMS_Za.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_12.clear()
        self.ui.plainTextEdit_12.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_23.clear()
        self.ui.plainTextEdit_23.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_29.clear()
        self.ui.plainTextEdit_29.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_28.clear()
        self.ui.plainTextEdit_28.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_21.clear()
        self.ui.plainTextEdit_21.setStyleSheet("background-color: white;")
        self.ui.DI1_H_3.clear()
        self.ui.DI1_H_3.setStyleSheet("background-color: white;")
        self.ui.IGN_H.clear()
        self.ui.IGN_H.setStyleSheet("background-color: white;")
        self.ui.DI1_H_6.clear()
        self.ui.DI1_H_6.setStyleSheet("background-color: white;")
        self.ui.DI1_H_7.clear()
        self.ui.DI1_H_7.setStyleSheet("background-color: white;")
        self.ui.DI1_H_4.clear()
        self.ui.DI1_H_4.setStyleSheet("background-color: white;")
        self.ui.DI1_H_5.clear()
        self.ui.DI1_H_5.setStyleSheet("background-color: white;")
        self.ui.DI1_H_8.clear()
        self.ui.DI1_H_8.setStyleSheet("background-color: white;")
        self.ui.DI_H.clear()
        self.ui.DI_H.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_24.clear()
        self.ui.plainTextEdit_24.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_22.clear()
        self.ui.plainTextEdit_22.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_30.clear()
        self.ui.plainTextEdit_30.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_26.clear()
        self.ui.plainTextEdit_26.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_25.clear()
        self.ui.plainTextEdit_25.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_31.clear()
        self.ui.plainTextEdit_31.setStyleSheet("background-color: white;")


        

    def check_barcode(self):
        # Get the barcode text from the input field
        self.ui.barcode_Input_2.setFocus()  # Set focus to the barcode input field
        time.sleep(0.1)
        #self.barcode = self.ui.barcode_Input_2.toPlainText().strip()  # Get text and remove any leading/trailing whitespaces
        #self.barcode = self.barcode.replace('\n', '').replace('\r', '').strip()
        print(f"Raw Barcode: {repr(self.barcode)}")
        print(f"Barcode: {self.barcode}")

        # Check if barcode text is valid (non-empty)
        if self.barcode:
            self.ui.barcode_Input_2.setPlainText(str(self.barcode))  # Copy the barcode to another input field

            # Barcode is received, make the API request
            # try:
            #     response = requests.put(self.stage4_url, json=self.data, headers=self.headers)
            #     print(response)
            #
            #     if response.status_code == 200:
            #         print('Barcode data scanned successfully')
            #         self.ui.pushButton_2.setEnabled(True)  # Enable the button if barcode data is valid
            #         self.check_previous_stages()  # Continue with previous stages
            #         self.ui.barcode_Input_2.setPlainText(self.barcode)  # Set actual barcode value
            #     else:
            #         print(f'Error reading barcode: {response.status_code}')
            #         self.ui.barcode_Input_2.setPlainText(
            #             'Error reading barcode')  # Set error message if barcode scan fails
            # except requests.exceptions.RequestException as e:
            #     print(f'API request failed: {e}')
            #     self.ui.barcode_Input_2.setPlainText('API request failed')  # Set API request failure message

    def on_button_click(self):
        # Call both actions (goToPage2 and login) in sequence
        print("Button clicked")

        self.login()
        self.showDateTime()
        self.check_server_status()
        self.check_barcode()
        if not self.timer.isActive():
            self.timer.start()  # Start the timer if not already active

    def showDateTime(self):
        # Get the current date and time
        self.current_datetime = QDateTime.currentDateTime().toString()

        # Display the date and time in the QPlainTextEdit widget
        self.ui.operator_Input_2.setPlainText(self.current_datetime)

    def show_message(self, title, message):
        # Show a message box to the user
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec_()

    def on_timer_timeout(self):
        # This method is called every time the timer times out (every second)
        print(f"Timer triggered. Elapsed time: {self.elapsed_time} seconds")

        # Increment the elapsed time by 1 second
        self.elapsed_time += 1

        hours, remainder = divmod(self.elapsed_time, 3600)  # Get hours and remainder
        minutes, seconds = divmod(remainder, 60)  # Get minutes and seconds

        # Format the time as hh:mm:ss
        formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"

        # Update the operator_Input_3 field with the elapsed time
        self.ui.operator_Input_3.setPlainText(f"{formatted_time}")

    def login(self):
        # Retrieve the text from the text fields
        self.operator = self.ui.plainTextEdit.toPlainText()  # Retrieve the operator input
        self.qc_head = self.ui.plainTextEdit_2.toPlainText()  # Retrieve the QC head input

        # Debugging print statements to check the retrieved values
        print("operator:", repr(self.operator))  # Using repr() to detect empty strings
        print("qc_head:", repr(self.qc_head))

        # Check if both fields are filled
        if self.operator == "" or self.qc_head == "":
            print("Either operator or qc_head is empty!")  # Debugging message
            self.show_message("Login Failed", "Please Enter operator and qc head")
            # Ensure navigation doesn't happen when either field is empty
        else:
            # If both fields are filled, proceed with the login success logic
            print("operator:", self.operator)
            print("qc_head:", self.qc_head)

            self.get_device_model()
            #self.show_message("Login Successful", "Welcome, Operator!")

            # Set the values in the corresponding UI input fields
            self.ui.operator_Input.setPlainText(self.operator)
            self.ui.QC_input.setPlainText(self.qc_head)

            # Navigate to Page 2 only if both fields are filled
            self.stackedWidget.setCurrentIndex(1)  # Go to Page 2

    def get_device_model(self):
        # Make the GET request to fetch device information based on barcode
        print("get_device_model")
        # Include the barcode in the payload (or query if required by the API)
        try:
            # Make the API request
            response = requests.put(self.stage5_url, json=self.data, headers=self.headers)
            print("get_device_model",response)
            if response.status_code == 200:
                # Parse the response JSON
                response_data = response.json()
                self.ui.barcode_Input_2.setPlainText(self.barcode)
                # Extract the model name from the response
                self.model_name = response_data.get("device", {}).get("model_name", "Model name not found")
                self.ui.QC_input_2.setPlainText(self.model_name)
                if self.model_name is not None:
                    self.select_parameters()
                # Display the model name in the UI or print it
                    print(f"Device Model: {self.model_name}")

                # If you want to show it in the UI, you can use something like:
                # self.ui.modelNameLabel.setText(model_name)

            else:
                print(f"Error fetching device data. Status code: {response.status_code}")

        except requests.exceptions.RequestException as e:
            # Handle any errors in the request
            print(f"Error: {e}")

    def goToPage2(self):
        print("Navigating to Page 2")
        self.stackedWidget.setCurrentIndex((self.stackedWidget.currentIndex() + 1) % 2)

    def check_server_status(self):
        response = requests.get("http://192.168.2.253:6101/api/PROD_check")
        if response.status_code == 200:
            print("response",response.status_code)
            self.ui.barcode_Input.setPlainText("Connected")
        else:
            self.ui.barcode_Input.setPlainText("Please check connection")

    def check_previous_stages(self):
        url = "http://192.168.2.253:6101/api/test_points/SENSOR001"

        try:
            # Send a GET request to the URL
            response = requests.get(url)

            # Check if the request was successful (status code 200)
            if response.status_code == 200:
                data = response.json()  # Parse the JSON response

                # Access the mechanical_fitting_status
                mechanical_fitting_status = data.get('device', {}).get('status', {}).get('mechanical_fitting_status',
                                                                                         False)
                print("mechanical_fitting_status", mechanical_fitting_status)
                # Check the status
                if mechanical_fitting_status:
                    print("Previous stage passed!")
                else:
                    print("Previous stage not passed.")
            else:
                print(f"Error: Unable to retrieve data, status code {response.status_code}")

        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")

    def select_parameters(self):
        print("select_parameters", self.model_name)
        if self.model_name == "ACON4L":

            self.ui.DI1_H_6.hide()
            self.ui.DI1_H_7.hide()
            self.ui.DI1_H_4.hide()
            self.ui.DI_H.hide()
            self.ui.DI1_H_5.hide()
            self.ui.DI1_H_8.hide()
            self.ui.DO1_L.hide()
            self.ui.DO1_H.hide()
            self.ui.DO2_L.hide()
            self.ui.DO2_H.hide()


            self.ui.label_50.hide()
            self.ui.label_49.hide()
            self.ui.label_40.hide()
            self.ui.label_41.hide()
            self.ui.label_41.hide()
            self.ui.label_42.hide()
        elif self.model_name == "Sampark AIS":
            self.ui.DI1_H_6.setDisabled(False)
            self.ui.DI1_H_7.setDisabled(False)
            self.ui.DI1_H_4.setDisabled(False)
            self.ui.DI_H.setDisabled(False)
            self.ui.DI1_H_5.setDisabled(False)
            self.ui.DI1_H_8.setDisabled(False)
            self.ui.DO1_L.setDisabled(False)
            self.ui.DO1_H.setDisabled(False)
            self.ui.DO2_L.setDisabled(False)
            self.ui.DO2_H.setDisabled(False)

    
    
    def send_data(self):
        print('button clicked')
        myobj = {"Status":True,
            "QR_code": "SENSOR965",  # This can be dynamically changed
    "visual_inspection_status": True,
    "visual_inspection_timestamp": {
        "$date": datetime.now(timezone.utc).isoformat()
    },
    "board_flashing_status": True,
    "board_flashing_timestamp": {
        "$date": datetime.now(timezone.utc).isoformat()
    },
    "screw_fitting_status": None,
    "mechanical_fitting_status": None,
    "mechanical_fitting_timestamp": None,
    "final_testing_status": None,
    "final_testing_timestamp": None,
    "IMEI": self.IMEI_ascii,
    "ICCID": "78910",
    "SystemRtc": None,
    "AppFWVersion": "88888",
    "BLFWVersion": None,
    "GPSFWVersion": None,
    "GSMFWVersion": None,
    "HWVersion": None,
    "GPSFix": None,
    "HDOP": None,
    "PDOP": 'xxxxx',
    "No_satelite": None,
    "GSMStatus": 'yyyyy',
    "signalStrength": None,
    "Network_code": None,
    "Network_Type": None,
    "SIM": None,
    "MEMS": None,
    "Voltage": None,
    "Memory": '111111',
    "Ignition": None,
    "Tamper": None,
    "DI_1_H": None,
    "DI_1_L": None,
    "DI_2_H": None,
    "DI_2_L": None,
    "DI_3_H": None,
    "DI_3_L": None,
    "DO_1_H": None,
    "DO_1_L": None,
    "DO_2_H": None,
    "DO_2_L": '22222',
    "CAN": None,
    "RS485": None,
    "AnalogInput1": None,
    "AnalogInput2": None,
    "sticker_printing_status": None,
    "sticker_printing_timestamp": None,
    "last_updated": {
        "$date": datetime.now(timezone.utc).isoformat()
    },
    "UID": None
        }
        headers = {'Content-Type': 'application/json'}  # Ensure correct headers
        x = requests.put(self.stage5_url, json=myobj, headers=headers,)
        print(x.status_code)

        if x.status_code == 200:
            print("POST request successful!")
            print(x.text)  # Print the response from the server
        else:
            print(f"POST request failed with status code {x.status_code}: {x.text}")


       
        

# Entry point of the program
if __name__ == "__main__":
    app = QApplication(sys.argv)    
    # Create an instance of the MyClass class
    processor = MyClass()
    processor.show()
    sys.exit(app.exec_())