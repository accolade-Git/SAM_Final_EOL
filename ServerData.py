# previous code with sejal ui
import can
import time
from PyQt5.QtCore import Qt
import binascii
import serial
import sys
import requests
import datetime
import pandas as pd
from datetime import datetime, timedelta, timezone
from finalTesting import Ui_FinalTestingUtility
import resources_rc
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox, QFileDialog, QSizePolicy
from PyQt5.QtCore import QTimer
from PyQt5.QtCore import QDateTime
from openpyxl import Workbook, load_workbook
from PyQt5.QtGui import QCursor
from PyQt5.QtGui import QPixmap
from PyQt5.QtGui import QIcon
from CAN_data_threading import Worker ,CAN_Data
from PyQt5.QtCore import QPoint
from openpyxl.styles import Alignment, Font, PatternFill
import logging
import os
from PyQt5.QtWidgets import QApplication, QWidget, QLabel
from PyQt5.QtCore import QThread, pyqtSignal
import threading

# from mainwindow import Ui_MainWindow
class TimerThread(QThread):
    # Signal to update the UI with the elapsed time
    update_time_signal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.elapsed_time = 0

    def run(self):
        while True:
            QThread.sleep(1)  # Sleep for 1 second
            self.elapsed_time += 1

            # Calculate hours, minutes, seconds
            hours, remainder = divmod(self.elapsed_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"

            # Emit the signal to update the UI
            self.update_time_signal.emit(formatted_time)


class MyClass(QMainWindow):
    def __init__(self,can,worker):
        super().__init__()
        self.ui = ui  # Use the passed ui (Ui_FinalTestingUtility instance)
        self.ui.setupUi(self)  # Setup UI
        self.CAN_data_obj = CAN_Data(self.ui)  # Pass the UI to CAN_Data
        self.worker_obj = Worker()  # No need to pass ui to Worker anymore

        self.ui.comboBox.addItem("Select Device Model")
        self.ui.comboBox.addItem("ACON4L")
        self.ui.comboBox.addItem("ACON4S")

        self.ui.comboBox.currentIndexChanged.connect(self.get_device_model)
        
        screen = QApplication.primaryScreen()
        screen_size = screen.size()

        # Load the image
        logo = QPixmap("AEPL_Logo.png")  # replace with your image file
        # Create a QLabel widget to display the image
        label = self.ui.label_19
        label.setPixmap(logo)

        # Set the window size based on screen resolution (100% of screen size)
        self.setGeometry(0, 0, screen_size.width(), screen_size.height())
        self.setMinimumSize(800, 600)
        self.stackedWidget = self.ui.stackedWidget
        self.selected_value = None
        self.stackedWidget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.barcode = None
        #self.ui.pushButton_2.setEnabled(True)
        self.stage4_url = "http://192.168.2.253:6101/api/stage4"
        self.stage5_url = "http://192.168.2.253:6101/api/stage5"
        self.device_status_url = f"http://192.168.2.253:6101/api/test_points/{self.barcode}"
        self.check_server_url = "http://192.168.2.253:6101/api/PROD_check"
        self.Barcode = None
        self.model_name = None
        # Initialize a QTimer
        self.timer_thread = TimerThread()
        self.timer_thread.start()  # Start the thread

        self.timer = QTimer(self)
        self.timer.setInterval(1000)  # 1000 ms = 1 second
        self.timer.timeout.connect(self.on_timer_timeout)
        self.elapsed_time = 0
        self.operator = None
        self.qc_head = None
        # Initialize CAN bus in the __init__ method to avoid reinitializing it every time
        cursor = QCursor()
        cursor.setPos(620, 70)
        self.totalParams = 36
        self.processed_params = 0
        self.sent_data = False
        self.UIN = None
        self.WDTUI_checkedin =0
        self.retry_flag =False

        self.satellite_count = None
        self.data = {
            "QR_code": self.barcode,
            "model_name": None,
            "visual_inspection_status": True,
            "visual_inspection_timestamp": None,
            "board_flashing_status": True,
            "board_flashing_timestamp": None,
            "screw_fitting_status": True,
            "screw_fitting_timestamp": None,
            "mechanical_fitting_status": True,
            "mechanical_fitting_timestamp": None,
            "final_testing_status": True,
            "final_testing_timestamp": None,
            "IMEI": None,
            "ICCID": None,
            "SystemRtc": None,
            "AppFWVersion": None,
            "BLFWVersion": None,
            "GPSFWVersion": None,
            "GSMFWVersion": None,
            "HWVersion": None,
            "GPSFix": None,
            "HDOP": None,
            "PDOP": None,
            "No_satelite": None,
            "GSMStatus": None,
            "signalStrength": None,
            "Network_code": None,
            "Network_Type": None,
            "SIM": None,
            "MEMS": None,
            "Voltage": None,
            "Memory": None,
            "Ignition": None,
            "Tamper": None,
            "DI_1_H": None,
            "DI_1_L": None,
            "DI_2_H": None,
            "DI_2_L": None,
            "DI_3_H": None,
            "DI_3_L": None,
            "DO_1_H": None,
            "DO_1_L": None,
            "DO_2_H": None,
            "DO_2_L": None,
            "CAN": None,
            "RS485": None,
            "AnalogInput1": None,
            "AnalogInput2": None,
            "sticker_printing_status": None,
            "sticker_printing_timestamp": datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f'),
            "last_updated": datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f'),
            "UIN": None
        }
        self.headers = {"Content-Type": "application/json"}
        #self.ui.pushButton_2.clicked.connect(self.save_to_excel)
        self.ui.pushButton.clicked.connect(self.on_button_click)
        self.ui.pushButton_2.clicked.connect(self.send_data)
        self.ui.pushButton_2.clicked.connect(self.save_to_excel)
        self.ui.pushButton_2.clicked.connect(self.CAN_data_obj.clear_ui)
        self.ui.pushButton_8.clicked.connect(self.delayed_start)
        self.ui.pushButton_9.clicked.connect(self.worker_obj.failed_func)

        self.worker_obj.update_103_singal.connect(self.updateUI_103)
        self.worker_obj.update_104_singal.connect(self.updateUI_104)
        self.worker_obj.update_106_singal.connect(self.updateUI_106)
        self.worker_obj.update_105_singal.connect(self.updateUI_105)
        self.worker_obj.update_101_singal.connect(self.updateUI_101)
        self.worker_obj.update_100_singal.connect(self.updateUI_100)
        self.worker_obj.update_110_singal.connect(self.updateUI_110)
        self.worker_obj.update_112_singal.connect(self.updateUI_112)
        self.worker_obj.update_109_singal.connect(self.updateUI_109)
        self.worker_obj.update_115_singal.connect(self.updateUI_115)
        self.worker_obj.update_116_singal.connect(self.updateUI_116)
        self.worker_obj.update_113_singal.connect(self.updateUI_113)
        self.worker_obj.update_114_singal.connect(self.updateUI_114)
        self.worker_obj.update_102_singal.connect(self.updateUI_102)
        self.worker_obj.update_121_singal.connect(self.updateUI_121)
        self.worker_obj.update_122_signal.connect(self.updateUI_122)
        self.worker_obj.update_123_singal.connect(self.updateUI_123)
        self.worker_obj.update_119_singal.connect(self.updateUI_119)
        self.worker_obj.update_retryUI_signal.connect(self.updateRetry_UI)

    def delayed_start(self):
        self.ui.plainTextEdit_12.setPlainText('Please wait device is initializing')
        # Create a QTimer for a 15-second delay
        timer = QTimer(self)
        timer.setSingleShot(True)  # Run the timer only once
        timer.timeout.connect(self.worker_obj.start_functions)
        timer.start(15000)  # 15000 milliseconds = 15 seconds
        

    def updateUI_103(self, app_ver):
        try:
            self.ui.pushButton_8.setEnabled(False)
            self.ui.pushButton_9.setEnabled(False)
            self.ui.pushButton_2.setEnabled(False)

            self.ui.plainTextEdit_8.setPlainText(app_ver)
            self.ui.plainTextEdit_12.setPlainText(f"Application Version : {app_ver}\n")

            if app_ver != 'SAM01_PROD_0.0.1_TST02':
                self.ui.plainTextEdit_8.setStyleSheet("background-color: red;")
            else:
                self.ui.plainTextEdit_8.setStyleSheet("background-color: white;")
        except Exception as e:
                # Log the exception or display an error message
                self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        finally:
            self.ui.plainTextEdit_8.setReadOnly(True)

    def updateUI_104(self, BL_ver):
        try:
            self.ui.plainTextEdit_9.setPlainText(BL_ver)
            self.ui.plainTextEdit_12.appendPlainText(f"BootLoader Version : {BL_ver}\n")

            if BL_ver != 'SAM01_BOOT_0.0.1_TST03':
                self.ui.plainTextEdit_9.setStyleSheet("background-color: red;")
                      
            else:
                self.ui.plainTextEdit_9.setStyleSheet("background-color: white;")
        except Exception as e:
                self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        finally:
            self.ui.plainTextEdit_9.setReadOnly(True)

    def updateUI_106(self,gsm_ver):
        try:
            self.ui.plainTextEdit_6.setPlainText(gsm_ver)
            self.ui.plainTextEdit_12.appendPlainText(f"GSM Version : {gsm_ver}\n")

            if gsm_ver != 'EC200UCNAAR03A11M08':
                self.ui.plainTextEdit_6.setStyleSheet("background-color: red;")
            else:
                self.ui.plainTextEdit_6.setStyleSheet("background-color: white;")
        except Exception as e:
                self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        finally:
            self.ui.plainTextEdit_6.setReadOnly(True)

    def updateUI_105(self,gps_ver):
        try:
            self.ui.plainTextEdit_5.setPlainText(gps_ver)
            self.ui.plainTextEdit_12.appendPlainText(f"GPS Version : {gps_ver}\n")      
            
            if gps_ver != 'EC200UCNAAR03A11M08':
                self.ui.plainTextEdit_5.setStyleSheet("background-color: red;")
            else:
                self.ui.plainTextEdit_5.setStyleSheet("background-color: white;")
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        finally:
            self.ui.plainTextEdit_5.setReadOnly(True)

    def updateUI_101(self,ICCID):
        try:
            self.ui.plainTextEdit_11.setPlainText(ICCID)
            self.ui.plainTextEdit_12.appendPlainText(f"ICCID : {ICCID}\n")
        
            if len(ICCID)<20:
                self.ui.plainTextEdit_11.setStyleSheet("background-color: red;")
            else:
                self.ui.plainTextEdit_11.setStyleSheet("background-color: white;") 

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")   
        finally:
            self.ui.plainTextEdit_11.setReadOnly(True)

    def updateUI_100(self,IMEI):
        try:
            self.ui.plainTextEdit_10.setPlainText(IMEI)
            self.ui.plainTextEdit_12.appendPlainText(f"IMEI : {IMEI}\n")
                    
            if len(IMEI) < 15:
                 self.ui.plainTextEdit_10.setStyleSheet("background-color: red;")
            else:
                 self.ui.plainTextEdit_10.setStyleSheet("background-color: white;")
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        finally:
            self.ui.plainTextEdit_10.setReadOnly(True)
            self.ui.plainTextEdit_12.setReadOnly(True)

    def updateUI_110(self,CREG,CGREG,CSQ):
        try:
            self.ui.CSQ.setPlainText(str(CREG))
            if CREG !=5:
                self.ui.CSQ.setStyleSheet("background-color: red;")
            else:
                self.ui.CSQ.setStyleSheet("background-color: white;")
                self.CREG_found = True

            self.ui.CGREG.setPlainText(str(CGREG))
            if CGREG != 5:
                self.ui.CGREG.setStyleSheet("background-color: red;")
            else:
                self.ui.CGREG.setStyleSheet("background-color: white;")
                self.CGREG_found =True

            self.ui.CREG.setPlainText(str(CSQ))
            if CSQ < 10:
                self.ui.CREG.setStyleSheet("background-color: red;")
            else:
                self.ui.CREG.setStyleSheet("background-color: white;")
                self.CSQ_found =True

            if self.CREG_found and self.CGREG_found:
                self.ui.plainTextEdit_24.setPlainText("Pass")
                self.ui.plainTextEdit_24.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                self.ui.plainTextEdit_24.setPlainText("Fail")
                self.ui.plainTextEdit_24.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.worker_obj.GSM_result = self.ui.plainTextEdit_24.toPlainText()

            self.ui.plainTextEdit_12.appendPlainText(f"CREG: {str(CREG)}\n")
            self.ui.plainTextEdit_12.appendPlainText(f"CGREG: {str(CGREG)}\n")
            self.ui.plainTextEdit_12.appendPlainText(f"CSQ : {str(CSQ)}\n")
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        finally:
            self.ui.CSQ.setReadOnly(True)
            self.ui.CGREG.setReadOnly(True)
            self.ui.CREG.setReadOnly(True)
            self.ui.plainTextEdit_24.setReadOnly(True)

    def updateUI_112(self,operator):
        try:
            self.ui.Operator.setPlainText(operator)
            self.ui.plainTextEdit_12.appendPlainText(f"Operator Name : {operator}\n")
                    
            if operator == 'AIRTEL 4G':
                self.ui.Operator.setStyleSheet("background-color: white;")
                self.operator_found = True
            else:
                self.ui.Operator.setStyleSheet("background-color: red;")
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
             self.ui.Operator.setReadOnly(True)

    def updateUI_109(self,gps_status,No_Of_Sat):
        try:
            self.ui.Operator_2.setPlainText(str(gps_status))
            self.ui.plainTextEdit_12.appendPlainText(f"GPS status: {str(gps_status)}\n")
           
            if gps_status != 1:
                self.ui.Operator_2.setStyleSheet("background-color:red")
                self.ui.plainTextEdit_28.setPlainText("Fail")
                self.ui.plainTextEdit_28.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:red""")
            else:
                self.ui.plainTextEdit_28.setPlainText("Pass")
                self.ui.plainTextEdit_28.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")

            self.worker_obj.Gps_result = self.ui.plainTextEdit_28.toPlainText()
            self.satellite_count = self.ui.NoOf_satellite.toPlainText()

            if No_Of_Sat ==3000 or No_Of_Sat == 184:
                self.ui.NoOf_satellite.setPlainText(str('0'))
                self.ui.NoOf_satellite.setStyleSheet("background-color: red;")
                self.ui.plainTextEdit_12.appendPlainText(f"No. of Satellite: {str('0')}\n")
            else:
                self.ui.NoOf_satellite.setPlainText(str(No_Of_Sat))
                self.ui.NoOf_satellite.setStyleSheet("background-color: white;")
                self.ui.plainTextEdit_12.appendPlainText(f"No. of Satellite: {str(No_Of_Sat)}\n")
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
            self.ui.Operator_2.setReadOnly(True)
            self.ui.NoOf_satellite.setReadOnly(True)
            self.ui.plainTextEdit_28.setReadOnly(True)

    def updateUI_115(self,mainsVtg : float):
        try:
            mainsVtg_str = f"{mainsVtg:.2f}"  # Format to 2 decimal places
            self.ui.mains_input_2.setPlainText(mainsVtg_str)
            self.ui.plainTextEdit_12.appendPlainText(f"Mains Voltage: {mainsVtg}\n")

            if 9.00 <= mainsVtg <= 32.00:
                      self.ui.mains_input_2.setStyleSheet("background-color: white;")
                      self.ui.plainTextEdit_29.setPlainText("Pass")
                      self.ui.plainTextEdit_29.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")
            else:
                      self.ui.mains_input_2.setStyleSheet("background-color: red;")
                      self.ui.plainTextEdit_29.setPlainText("Fail")
                      self.ui.plainTextEdit_29.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:red""")

            self.worker_obj.Mains_result = self.ui.plainTextEdit_29.toPlainText()

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
            self.ui.mains_input_2.setReadOnly(True)
            self.ui.plainTextEdit_29.setReadOnly(True)

    def updateUI_116(self,IntBatVtg):
        try:
            IntBatVtg_float = float(IntBatVtg)
            self.ui.IntBat_input_2.setPlainText(IntBatVtg)
            self.ui.plainTextEdit_12.appendPlainText(f"Int_Bat Voltage: {IntBatVtg}\n")

            if IntBatVtg_float > 4.2 or IntBatVtg_float == 0.00:
                    self.ui.IntBat_input_2.setStyleSheet("background-color: red;")
                    self.ui.plainTextEdit_23.setPlainText("Fail")
                    self.ui.plainTextEdit_23.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:red""")
            else:
                    self.ui.IntBat_input_2.setStyleSheet("background-color: white;")
                    self.ui.plainTextEdit_23.setPlainText("Pass")
                    self.ui.plainTextEdit_23.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")

            self.worker_obj.IntVtg_result = self.ui.plainTextEdit_23.toPlainText()

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")
        
        finally:
            self.ui.IntBat_input_2.setReadOnly(True)
            self.ui.plainTextEdit_23.setReadOnly(True)

    def updateUI_113(self,Mqtt_status,LoginPacket):
        try:
            if Mqtt_status == 0 or Mqtt_status == 255:
                    self.ui.Analog1_2.setStyleSheet("background-color: red;")
                    self.ui.Analog1_2.setPlainText(str('0'))
                    self.ui.plainTextEdit_21.setPlainText("Fail")
                    self.ui.plainTextEdit_21.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
            else:
                    self.ui.Analog1_2.setStyleSheet("background-color: white;")
                    self.ui.Analog1_2.setPlainText(str(Mqtt_status))
                    self.ui.plainTextEdit_21.setPlainText("Pass")
                    self.ui.plainTextEdit_21.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

            self.worker_obj.MQTT_result = self.ui.plainTextEdit_21.toPlainText()

            self.ui.MQTT.setPlainText(str(LoginPacket))
            if LoginPacket != 1:
                self.ui.MQTT.setStyleSheet("background-color: red;")
            else:
                self.ui.MQTT.setStyleSheet("background-color: white;")

            self.ui.plainTextEdit_12.appendPlainText(f"MQTT status: {str(LoginPacket)}\n")
            self.ui.plainTextEdit_12.appendPlainText(f"No. of Login packet: {str(LoginPacket)}\n")
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
            self.ui.Analog1_2.setReadOnly(True)
            self.ui.MQTT.setReadOnly(True)
            self.ui.plainTextEdit_31.setReadOnly(True)

    def updateUI_114(self,frame1,frame2,frame3,prefix):
        try:
            self.ui.MEMS_Xa.setPlainText(frame1)
            self.ui.MEMS_Ya.setPlainText(frame2)
            self.ui.MEMS_Za.setPlainText(frame3)
            self.ui.MEMS_Xa_2.setPlainText(prefix)
            self.ui.plainTextEdit_12.appendPlainText(f"Accelerometer data: {frame1}, {frame2}, {frame3}")

            # Determine pass/fail status based on the frames
            if frame1 and frame2 and frame3:
                self.ui.plainTextEdit_25.setPlainText("Pass")
                self.ui.plainTextEdit_25.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                self.ui.plainTextEdit_25.setPlainText("Fail")
                self.ui.plainTextEdit_25.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.worker_obj.MEMS_result = self.ui.plainTextEdit_25.toPlainText()

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

           
        finally:
            self.ui.MEMS_Xa.setReadOnly(True)
            self.ui.MEMS_Ya.setReadOnly(True)
            self.ui.MEMS_Za.setReadOnly(True)
            self.ui.plainTextEdit_25.setReadOnly(True)

    def updateUI_102(self,RTC):
        try:
            if RTC.total_seconds()>15:
                    self.ui.plainTextEdit_13.setPlainText("Fail")
                    self.ui.plainTextEdit_13.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
            else:
                    self.ui.plainTextEdit_13.setPlainText("Pass")
                    self.ui.plainTextEdit_13.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

            self.worker_obj.RTC_result = self.ui.plainTextEdit_13.toPlainText()

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

            
        finally:
            self.ui.plainTextEdit_21.setReadOnly(True)

    def updateUI_121(self,device_id_flag, erase_status_flag, read_status_flag, write_status_flag,device_id,erase_status,read_status,write_status):
        try:
            if device_id_flag and erase_status_flag and read_status_flag and write_status_flag:
                    self.ui.plainTextEdit_51.setPlainText("Pass")
                    self.ui.plainTextEdit_51.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                    self.ui.plainTextEdit_51.setPlainText("Fail")
                    self.ui.plainTextEdit_51.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.worker_obj.Flash_result = self.ui.plainTextEdit_51.toPlainText()

            self.ui.plainTextEdit_12.appendPlainText(f"Device id: {str(device_id)}\n")
            self.ui.plainTextEdit_12.appendPlainText(f"Flash erase status: {str(erase_status)}\n")
            self.ui.plainTextEdit_12.appendPlainText(f"Flash read status : {str(read_status)}\n")
            self.ui.plainTextEdit_12.appendPlainText(f"Flash write status : {str(write_status)}\n")

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
            self.ui.plainTextEdit_51.setReadOnly(True)

    def updateUI_122(self,string1:str):
        try:
            if string1:
                self.ui.plainTextEdit_12.setPlainText(string1)
                self.ui.plainTextEdit_12.setStyleSheet("""
                    font-size: 16px; 
                    font-weight: bold; 
                    color: red;
                """)
        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")


    def updateUI_123(self,watchdogreboot_flag,watchdogrebootCount_flag,current_WDTCount):
        try:
            #self.WDTUI_checkedin +=1
            if watchdogreboot_flag and watchdogrebootCount_flag:
                        self.ui.plainTextEdit_12.setPlainText(f'Device reboot successful. \nCurrent bootcount is: {current_WDTCount}')
                        self.ui.plainTextEdit_14.setPlainText("Pass")
                        self.ui.plainTextEdit_14.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
            else:
                        self.ui.plainTextEdit_12.setPlainText(f'Device reboot successful. But timeout to get bootcount')
                        self.ui.plainTextEdit_14.setPlainText("Fail")
                        self.ui.plainTextEdit_14.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.worker_obj.WDT_result = self.ui.plainTextEdit_14.toPlainText()
           

            if self.worker_obj.Mains_result == 'Pass' and self.worker_obj.IntVtg_result == 'Pass' and self.worker_obj.Gps_result == 'Pass' and self.worker_obj.GSM_result == 'Pass' and self.worker_obj.IGN_result and self.worker_obj.Tamper_result and  self.worker_obj.RTC_result == 'Pass' and self.worker_obj.MEMS_result == 'Pass' and self.worker_obj.MQTT_result == 'Pass'and self.worker_obj.WDT_result == 'Pass':
                    self.update_overallResult()
                
            else:
                self.ui.textEdit.setPlainText('Pending')
                #self.update_overallResult()
                self.ui.pushButton_9.setEnabled(True)
                self.ui.pushButton_8.setEnabled(True)
                self.ui.pushButton_2.setEnabled(True)

                

            # if self.WDTUI_checkedin ==1:
            #     self.ui.pushButton_9.setEnabled(True)
            #     self.ui.pushButton_8.setEnabled(True)
            #     self.ui.pushButton_2.setEnabled(True)
            # else:
            #     pass

            # self.worker_obj.overall_result = self.ui.textEdit.toPlainText()

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
            self.ui.plainTextEdit_14.setReadOnly(True)
            #self.ui.pushButton_2.setEnabled(True) 

    def updateUI_119(self,IGN,tamper):
        try:
            if IGN == 1:
                self.ui.plainTextEdit_12.appendPlainText(f'IGN status: {IGN}')
                self.ui.plainTextEdit_26.setPlainText("Pass")
                self.ui.plainTextEdit_26.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

            else:
                self.ui.plainTextEdit_12.appendPlainText(f'IGN status: {IGN}')
                self.ui.plainTextEdit_26.setPlainText("Fail")
                self.ui.plainTextEdit_26.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.CAN_data_obj.IGN_result = self.ui.plainTextEdit_26.toPlainText()

            if tamper == 0:
                self.ui.plainTextEdit_12.appendPlainText(f'Tamper status: {tamper}')
                self.ui.plainTextEdit_47.setPlainText("Pass")
                self.ui.plainTextEdit_47.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

            else:
                self.ui.plainTextEdit_12.appendPlainText(f'Tamper status: {tamper}')
                self.ui.plainTextEdit_47.setPlainText("Fail")
                self.ui.plainTextEdit_47.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")

            self.worker_obj.Tamper_result = self.ui.plainTextEdit_47.toPlainText()

        except Exception as e:
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")

        finally:
            self.ui.plainTextEdit_26.setReadOnly(True)
            self.ui.plainTextEdit_47.setReadOnly(True) 

    def updateRetry_UI(self, string1: str, flag1: bool):  # Ensure flag1 is a boolean
        try:
            self.retry_flag = flag1
            print(f"Debug: string1={string1}, flag1={flag1}")  # Log values for debugging
            
            # Check if string1 is valid
            if string1:
                self.ui.plainTextEdit_12.setPlainText(string1)
                self.ui.plainTextEdit_12.setStyleSheet("""
                    font-size: 16px; 
                    font-weight: bold; 
                    color: red;
                """)
                # Disable buttons
                self.ui.pushButton_9.setEnabled(False)
                self.ui.pushButton_8.setEnabled(False)
                self.ui.pushButton_2.setEnabled(False)
                print("Buttons disabled.")

            # Check if flag1 is True
            if flag1:
                string2 = 'Max retries done. Final Testing done.'
                self.ui.plainTextEdit_12.setStyleSheet("""
                    font-size: 20px; 
                    font-weight: bold; 
                    color: green;
                """)
                self.ui.plainTextEdit_12.setPlainText(string2)
                self.timer.stop()
                self.update_overallResult()
                # self.ui.pushButton_9.setEnabled(True)
                # self.ui.pushButton_8.setEnabled(True)
                # self.ui.pushButton_2.setEnabled(True)

            #Overall result check
            # if (
            #     self.worker_obj.Mains_result == 'Pass'
            #     and self.worker_obj.IntVtg_result == 'Pass'
            #     and self.worker_obj.Gps_result == 'Pass'
            #     and self.worker_obj.GSM_result == 'Pass'
            #     and self.worker_obj.IGN_result == 'Pass'
            #     and self.worker_obj.Tamper_result == 'Pass'
            #     and self.worker_obj.RTC_result == 'Pass'
            #     and self.worker_obj.MEMS_result == 'Pass'
            #     and self.worker_obj.MQTT_result == 'Pass'
            #     and self.worker_obj.WDT_result == 'Pass'
            # ):
            #     self.ui.textEdit.setText("Pass")
            #     self.ui.textEdit.setAlignment(Qt.AlignCenter)
            #     self.ui.textEdit.setStyleSheet("""
            #         font-size: 24px; 
            #         font-weight: bold; 
            #         background-color: green;
            #     """)
            # else:
            #     self.ui.textEdit.setText("Fail")
            #     self.ui.textEdit.setAlignment(Qt.AlignCenter)
            #     self.ui.textEdit.setStyleSheet("""
            #         font-size: 24px; 
            #         font-weight: bold; 
            #         background-color: red;
            #     """)

            # # Update overall result
            # self.worker_obj.overall_result = self.ui.textEdit.toPlainText()

          

        except Exception as e:
            print(f"Error in updateRetry_UI: {e}")  # Debugging output
            self.ui.plainTextEdit_12.appendPlainText(f"Error updating UI: {str(e)}\n")


    def update_overallResult(self):
        try:
            
                if self.worker_obj.Mains_result == 'Pass' and self.worker_obj.IntVtg_result == 'Pass' and self.worker_obj.Gps_result == 'Pass' and self.worker_obj.GSM_result == 'Pass' and self.worker_obj.IGN_result and self.worker_obj.Tamper_result and  self.worker_obj.RTC_result == 'Pass' and self.worker_obj.MEMS_result == 'Pass' and self.worker_obj.MQTT_result == 'Pass'and self.worker_obj.WDT_result == 'Pass':

                    self.ui.textEdit.setText("Pass")
                    self.ui.textEdit.setAlignment(Qt.AlignCenter)
                    self.ui.textEdit.setStyleSheet("""Font-size:24px; font-weight: Bold; background-color: green""")
                else:
                    self.ui.textEdit.setText("Fail")
                    self.ui.textEdit.setAlignment(Qt.AlignCenter)
                    self.ui.textEdit.setStyleSheet("""Font-size:24px; font-weight: Bold; background-color: red""")

                # Update overall result
                self.worker_obj.overall_result = self.ui.textEdit.toPlainText()

                self.ui.pushButton_9.setEnabled(True)
                self.ui.pushButton_8.setEnabled(True)
                self.ui.pushButton_2.setEnabled(True)


            
        except Exception as e:
                print(f"Error : {e}")
                
                

    def clean_string(self,input_string):
    
            # Filter out characters with ASCII values from 0-31 and 127
            return ''.join(c for c in input_string if 31 < ord(c) < 127)

    def save_to_excel(self):
        # Get the current date in YYYY-MM-DD format
        current_date = datetime.now().strftime("%Y-%m-%d")

        # Create a file path with the date appended to the filename
        current_directory = os.getcwd()
        file_path = os.path.join(current_directory, f"Final_Testing_DeviceStatus_{current_date}.xlsx")

        # If the file exists, load the existing workbook; otherwise, create a new one
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

            # Set the headers for the columns (only if the file is being created for the first time)
            headers = ['Date', 'IMEI', 'ICCID', 'Application Version','BL version', 'GSM Version',
                   'GPS Version', 'Mains vtg', 'Int_Bat vtg', 'GPS status', 'No.of Sat', 'CREG', 'CGREG', 'CSQ',
                   'Operator', 'MQTT', 'No.Of Login packet', 'MEMS Xa', 'MEMS Ya', 'MEMS Za', 'Mains result', 'IntBat result'
                   ,'Gps result', 'GSM result','IGN result','Tamper result','FlashMemory result','MEMS result','MQTT result','RTC result','WDT result','Overall Result']
            ws.append(headers)  # Append headers as the first row

        self.current_datetime =self.ui.operator_Input_2.toPlainText()

    

    # Clean the data before inserting into the worksheet
        data = [
        self.clean_string(str(self.current_datetime)) if self.current_datetime is not None else 'Not found',
        self.clean_string(self.worker_obj.IMEI_ascii) if self.worker_obj.IMEI_ascii is not None else 'Not found',
        self.clean_string(self.worker_obj.ICCID_ascii) if self.worker_obj.ICCID_ascii is not None else 'Not found',
        self.clean_string(self.worker_obj.appln_ver) if self.worker_obj.appln_ver is not None else 'Not found',
        self.clean_string(self.worker_obj.BL_ver) if self.worker_obj.BL_ver is not None else 'Not found',
        self.clean_string(self.worker_obj.GSM_ver) if self.worker_obj.GSM_ver is not None else 'Not found ',
        self.clean_string(self.worker_obj.Gps_ver) if self.worker_obj.Gps_ver is not None else 'Not found',
        self.clean_string(self.worker_obj.mains_vtg) if self.worker_obj.mains_vtg is not None else 'Not found',
        self.clean_string(self.worker_obj.Int_vtg) if self.worker_obj.Int_vtg is not None else 'Not found',
        self.clean_string(str(self.worker_obj.Gps_status)) if self.worker_obj.Gps_status is not None else 'Not found',
        self.clean_string(str(self.satellite_count)) if self.satellite_count is not None else 'Not found',
        self.clean_string(str(self.worker_obj.CREG)) if self.worker_obj.CREG is not None else 'Not found',
        self.clean_string(str(self.worker_obj.CGREG)) if self.worker_obj.CGREG is not None else 'Not found',
        self.clean_string(str(self.worker_obj.CSQ)) if self.worker_obj.CSQ is not None else 'Not found',
        self.clean_string(str(self.worker_obj.operatorName)) if self.worker_obj.operatorName is not None else 'Not found',
        self.clean_string(str(self.worker_obj.MQTT_status)) if self.worker_obj.MQTT_status is not None else 'Not found',
        self.clean_string(str(self.worker_obj.No_of_LogInPacket)) if self.worker_obj.No_of_LogInPacket is not None else 'Not found',
        self.clean_string(str(self.worker_obj.frame1)) if self.worker_obj.frame1 is not None else 'Not found',
        self.clean_string(str(self.worker_obj.frame2)) if self.worker_obj.frame2 is not None else 'Not found',
        self.clean_string(str(self.worker_obj.frame3)) if self.worker_obj.frame3 is not None else 'Not found',
        self.clean_string(str(self.worker_obj.Mains_result)) if self.worker_obj.Mains_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.IntVtg_result)) if self.worker_obj.IntVtg_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.Gps_result)) if self.worker_obj.Gps_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.GSM_result)) if self.worker_obj.GSM_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.IGN_result)) if self.worker_obj.IGN_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.Tamper_result)) if self.worker_obj.Tamper_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.Flash_result)) if self.worker_obj.Flash_result is not None else 'Not found',
        #self.clean_string(str(self.CAN_data_obj.DIs_result)) if self.CAN_data_obj.DIs_result is not None else 'Not found',
        #self.clean_string(str(self.CAN_data_obj.Dos_result)) if self.CAN_data_obj.DOs_result is not None else 'Not found',
        #self.clean_string(str(self.CAN_data_obj.AnalogVolt_result)) if self.CAN_data_obj.AnalogVolt_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.MEMS_result)) if self.worker_obj.MEMS_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.MQTT_result)) if self.worker_obj.MQTT_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.RTC_result)) if self.worker_obj.RTC_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.WDT_result)) if self.worker_obj.WDT_result is not None else 'Not found',
        self.clean_string(str(self.worker_obj.overall_result)) if self.worker_obj.overall_result is not None else 'Not found'
        ]

        # Append the data to the next available row
        ws.append(data)

        try:
            # Save the workbook to the specified path
            wb.save(file_path)
            print(f"Data saved to {file_path}")
            QMessageBox.information(self, "Success", f"Data successfully saved to {file_path}")

            # Clear the UI inputs after saving data
            #self.clear_ui()

        except Exception as e:
            print(f"Error saving data to Excel: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to save data to Excel: {str(e)}")

    
    def on_timer_timeout(self):
        # This method is called every time the timer times out (every second)
 
        # Increment the elapsed time by 1 second
        self.elapsed_time += 1
 
        hours, remainder = divmod(self.elapsed_time, 3600)  # Get hours and remainder
        minutes, seconds = divmod(remainder, 60)  # Get minutes and seconds
 
        # Format the time as hh:mm:ss
        formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"
 
        # Update the operator_Input_3 field with the elapsed time
        self.ui.operator_Input_3.setPlainText(f"{formatted_time}")

            
        
    def on_button_click(self):
        # Call both actions (goToPage2 and login) in sequence
        self.login()
        self.showDateTime()
        self.check_server_status()
        self.check_barcode()

    
    def check_barcode(self):
        # Clear the barcode input field before scanning (optional)
        self.ui.barcode_Input_2.clear()

        # Set focus to the barcode input field
        self.ui.barcode_Input_2.setFocus()

        # Define the function to check the barcode input
        def check_barcode_value():
            # Get the barcode value and strip leading whitespaces
            barcode = self.ui.barcode_Input_2.toPlainText().lstrip()
             
            # If the barcode is not empty, process it
            if barcode:
                self.barcode = barcode

                # Clean the barcode text if necessary
                self.barcode = self.barcode.strip()

                # Place the cleaned barcode back into the input field
                self.ui.barcode_Input_2.setPlainText(self.barcode)
                if not self.timer.isActive():
                    self.timer.start()  # Start the timer if not already active

                self.device_status_url = f"http://192.168.2.253:6101/api/test_points/{self.barcode}"

                #self.get_device_model()

                # Reset the cursor position to the beginning of the input field
                cursor = self.ui.barcode_Input_2.textCursor()
                cursor.setPosition(0)
                self.ui.barcode_Input_2.setTextCursor(cursor)

                # Stop the timer once the barcode is processed
                timer.stop()  # This stops the timer inside the function

                # Proceed to the next stage or action
                self.check_previous_stages()

        # Set up a QTimer to check the input field every 100 milliseconds
        # Ensure the timer is started only once and stops when finished
        timer = QTimer(self)
        timer.timeout.connect(check_barcode_value)

        # Start the timer to check every 100 milliseconds
        timer.start(100)  # Check every 100 milliseconds


    def showDateTime(self):
        # Get the current date and time
        current_datetime = QDateTime.currentDateTime().toString()

        # Display the date and time in the QPlainTextEdit widget
        self.ui.operator_Input_2.setPlainText(current_datetime)


    def show_message(self, title, message):
        # Show a message box to the user
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec_()

    def login(self):
        # Retrieve the text from the text fields
        self.operator = self.ui.plainTextEdit.toPlainText()  # Retrieve the operator input
        self.qc_head = self.ui.plainTextEdit_2.toPlainText()  # Retrieve the QC head input

        self.model_name = self.ui.comboBox.currentText()

        # Debugging print statements to check the retrieved values
        # print("operator:", repr(self.operator))  # Using repr() to detect empty strings
        # print("qc_head:", repr(self.qc_head))

        # Check if both fields are filled
        if self.operator == "" or self.qc_head == "":
            # print("Either operator or qc_head is empty!")  # Debugging message
            # self.show_message("Login Failed", "Please Enter operator and qc head")
            # Ensure navigation doesn't happen when either field is empty
            pass
        else:
            # If both fields are filled, proceed with the login success logic
            # print("operator:", self.operator)
            # print("qc_head:", self.qc_head)

            # self.show_message("Login Successful", "Welcome, Operator!")

            # Set the values in the corresponding UI input fields
            self.ui.operator_Input.setPlainText(self.operator)
            self.ui.QC_input.setPlainText(self.qc_head)

            # Navigate to Page 2 only if both fields are filled
            self.stackedWidget.setCurrentIndex(1)  # Go to Page 2


    def update_elapsed_time(self, formatted_time):
        # This method will be connected to the timer's signal and update the UI
        # Update the UI with the formatted time (this can be customized as needed)
        self.ui.operator_Input_3.setPlainText(formatted_time)

    def get_device_model(self):
        
        self.selected_value = self.ui.comboBox.currentText()
        if self.selected_value != "Select Device Model":
            # Assign the selected value to model_name and call get_device_model
            self.model_name = self.selected_value
            
        # Include the barcode in the payload (or query if required by the API)
        try:
            print(f"Getting device model: {self.model_name}")
            

            if self.model_name and self.model_name != "Select Device Model":
                print('model name',self.model_name)
                self.select_parameters()

                self.timer.start(1000)  # Check every 100 milliseconds
 
            else:
                # self.ui.QC_input_2.setPlainText("Model name not found")
                # self.ui.QC_input_2.setStyleSheet("background-color: red;")
                #self.ui.pushButton_8.setEnabled(False)
                self.ui.label_34.setEnabled(False)
                self.ui.label_13.setEnabled(False)
                self.ui.label_14.setEnabled(False)
                self.ui.label_15.setEnabled(False)
                self.ui.label_16.setEnabled(False)
                self.ui.label_43.setEnabled(False)
                self.ui.plainTextEdit_5.setEnabled(False)
                self.ui.plainTextEdit_6.setEnabled(False)
                self.ui.plainTextEdit_8.setEnabled(False)
                self.ui.plainTextEdit_9.setEnabled(False)
                self.ui.label_17.setEnabled(False)
                self.ui.label_18.setEnabled(False)
                self.ui.label_28.setEnabled(False)
                self.ui.label_29.setEnabled(False)
                self.ui.label_23.setEnabled(False)
                self.ui.label_38.setEnabled(False)
                self.ui.label_30.setEnabled(False)
                self.ui.label_31.setEnabled(False)
                self.ui.label_48.setEnabled(False)
                self.ui.label_51.setEnabled(False)
                self.ui.plainTextEdit_5.setEnabled(False)
                self.ui.plainTextEdit_6.setEnabled(False)
                self.ui.plainTextEdit_8.setEnabled(False)
                self.ui.plainTextEdit_9.setEnabled(False)
                self.ui.plainTextEdit_10.setEnabled(False)
                self.ui.Tamp_L.setEnabled(False)
                self.ui.label_54.setEnabled(False)
                self.ui.label_45.setEnabled(False)
                self.ui.label_46.setEnabled(False)
                self.ui.label_47.setEnabled(False)
                #self.ui.MEMS_INIT.setEnabled(False)
                self.ui.MEMS_Xa.setEnabled(False)
                self.ui.MEMS_Ya.setEnabled(False)
                self.ui.MEMS_Za.setEnabled(False)
                self.ui.label_53.setEnabled(False)
                self.ui.label_27.setEnabled(False)
                self.ui.label_35.setEnabled(False)
                self.ui.label_36.setEnabled(False)
                self.ui.label_32.setEnabled(False)
                self.ui.label_37.setEnabled(False)
                self.ui.mains_input_2.setEnabled(False)
                self.ui.IntBat_input_2.setEnabled(False)
                self.ui.Analog1_2.setEnabled(False)
                self.ui.Analog2_4.setEnabled(False)
                self.ui.Analog2_5.setEnabled(False)
                self.ui.Analog2_3.setEnabled(False)
                self.ui.Analog2_6.setEnabled(False)
                self.ui.label_39.setEnabled(False)
                self.ui.label_50.setEnabled(False)
                self.ui.label_49.setEnabled(False)
                self.ui.label_40.setEnabled(False)
                self.ui.label_41.setEnabled(False)
                self.ui.label_42.setEnabled(False)
                self.ui.label_39.setEnabled(False)
                self.ui.DI1_H_3.setEnabled(False)
                self.ui.IGN_H.setEnabled(False)
                self.ui.DI1_H_6.setEnabled(False)
                self.ui.DI1_H_7.setEnabled(False)
                self.ui.DI1_H_4.setEnabled(False)
                self.ui.DI1_H_5.setEnabled(False)
                self.ui.DI1_H_8.setEnabled(False)
                self.ui.DI_H.setEnabled(False)
                self.ui.DO1_L.setEnabled(False)
                self.ui.DO1_H.setEnabled(False)
                self.ui.DO2_L.setEnabled(False)
                self.ui.DO2_H.setEnabled(False)
                self.ui.label_55.setEnabled(False)
                
        finally:
            pass


    def goToPage2(self):
        # print("Navigating to Page 2")
        self.stackedWidget.setCurrentIndex((self.stackedWidget.currentIndex() + 1) % 2)


    def check_server_status(self):
        response = requests.get("http://192.168.2.253:6101/api/PROD_check")
        if response.status_code == 200:
            # print("response",response.status_code)
            self.ui.barcode_Input.setPlainText("Connected")
        else:
            self.ui.barcode_Input.setPlainText("Please check connection")


    def check_previous_stages(self):
        url = f"http://192.168.2.253:6101/api/test_points/{self.barcode}"
        # print("url",url)
        try:
            # Send a GET request to the URL
            response = requests.get(url)

            # Check if the request was successful (status code 200)
            if response.status_code == 200:
                data = response.json()  # Parse the JSON response

                # Access the mechanical_fitting_status
                mechanical_fitting_status = data.get('device', {}).get('status', {}).get('mechanical_fitting_status',
                                                                                         False)
                # print("mechanical_fitting_status", mechanical_fitting_status)
                # Check the status
                if mechanical_fitting_status:
                    self.ui.server_Input.setPlainText("Previous stage passed!")
                    self.ui.server_Input.setStyleSheet("background-color: green;")
                    #self.send_data()
                    #    print("Previous stage passed!")
                    return True
            else:

                self.ui.server_Input.setPlainText("Previous stage not passed.")  # Set the text
                self.ui.server_Input.setStyleSheet("background-color: red;")
                return False

        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")

        


        


    def select_parameters(self):
        print('inside select parameters')
        if self.model_name == "ACON4L":
            self.ui.label_32.hide()
            self.ui.Analog2_4.hide()
            self.ui.Analog2_5.hide()
            self.ui.label_37.hide()
            self.ui.Analog2_3.hide()
            self.ui.Analog2_6.hide()
            # self.ui.label_39.hide()
            # self.ui.DI1_H_3.hide()
            # self.ui.IGN_H.hide()
            self.ui.label_50.hide()
            self.ui.DI1_H_6.hide()
            self.ui.DI1_H_7.hide()
            self.ui.label_49.hide()
            self.ui.DI1_H_4.hide()
            self.ui.DI1_H_5.hide()
            self.ui.label_40.hide()
            self.ui.DI1_H_8.hide()
            self.ui.DI_H.hide()
            self.ui.label_41.hide()
            self.ui.DO1_L.hide()
            self.ui.DO1_H.hide()
            self.ui.label_42.hide()
            self.ui.DO2_L.hide()
            self.ui.DO2_H.hide()
            self.ui.label_73.hide()
            self.ui.plainTextEdit_22.hide()
            self.ui.label_64.hide()
            self.ui.plainTextEdit_27.hide()
            self.ui.label_75.hide()
            self.ui.plainTextEdit_31.hide()



        elif self.model_name == "ACON4S":
            self.ui.label_32.show()
            self.ui.Analog2_4.show()
            self.ui.Analog2_5.show()
            self.ui.label_37.show()
            self.ui.Analog2_3.show()
            self.ui.Analog2_6.show()
            # self.ui.label_39.show()
            # self.ui.DI1_H_3.show()
            # self.ui.IGN_H.show()
            self.ui.label_50.show()
            self.ui.DI1_H_6.show()
            self.ui.DI1_H_7.show()
            self.ui.label_49.show()
            self.ui.DI1_H_4.show()
            self.ui.DI1_H_5.show()
            self.ui.label_40.show()
            self.ui.DI1_H_8.show()
            self.ui.DI_H.show()
            self.ui.label_41.show()
            self.ui.DO1_L.show()
            self.ui.DO1_H.show()
            self.ui.label_42.show()
            self.ui.DO2_L.show()
            self.ui.DO2_H.show()
            self.ui.label_73.show()
            self.ui.plainTextEdit_22.show()
            self.ui.label_64.show()
            self.ui.plainTextEdit_27.show()
            self.ui.label_75.show()
            self.ui.plainTextEdit_31.show()

    def send_data(self):  # Update variable here to send to server
        print("Sending data...")

        if self.check_previous_stages and not self.sent_data:  # Check if data has not been sent yet
            try:
                self.data["QR_code"] = self.barcode
                self.data['model_name'] = self.model_name
                self.data["visual_inspection_timestamp"] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
                self.data["board_flashing_timestamp"] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
                self.data["screw_fitting_timestamp"] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
                self.data["mechanical_fitting_timestamp"] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
                self.data['IMEI'] = self.worker_obj.IMEI_ascii
                self.data['ICCID'] = self.worker_obj.ICCID_ascii
                self.data['SystemRtc'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'
                self.data['AppFWVersion'] = self.worker_obj.appln_ver
                self.data['BLFWVersion'] = None
                self.data['GPSFWVersion'] = self.worker_obj.Gps_ver
                self.data['GSMFWVersion'] = self.worker_obj.GSM_ver
                self.data['HWVersion'] = None
                self.data['GPSFix'] = self.worker_obj.Gps_status
                self.data['HDOP'] = None
                self.data['PDOP'] = None
                self.data['No_satelite'] = self.worker_obj.concatenated_satellites_decimal
                self.data['GSMStatus'] = self.worker_obj.GSM_result
                self.data['signalStrength'] = self.worker_obj.CSQ,
                self.data['Network_code'] = None
                self.data['Network_Type'] = self.worker_obj.operatorName
                self.data['SIM'] = None
                self.data['MEMS'] = self.worker_obj.MEMS_result
                self.data['Voltage'] = self.worker_obj.Mains_result
                self.data['Memory'] = None
                self.data['Ignition'] = self.worker_obj.IGN
                self.data['Tamper'] = self.worker_obj.tamper
                self.data['DI_1_H'] = self.worker_obj.DI1_seen_1
                self.data['DI_1_L'] = self.worker_obj.DI1_seen_0
                self.data['DI_2_H'] = self.worker_obj.DI2_seen_1
                self.data['DI_2_L'] = self.worker_obj.DI2_seen_0
                self.data['DI_3_H'] = self.worker_obj.DI3_seen_1
                self.data['DI_3_L'] = self.worker_obj.DI3_seen_0
                self.data['DO_1_H'] = None
                self.data['DO_1_L'] = None
                self.data['DO_2_H'] = None
                self.data['DI_2_L'] = None
                self.data['CAN'] = None
                self.data['RS485'] = None
                self.data['AnalogInput1'] = None
                self.data['AnalogInput2'] = None
                self.data['sticker_printing_status'] = None
                self.data['sticker_printing_timestamp'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')
                self.data['last_updated'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')
                self.data['sticker_printing_timestamp'] = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f')

                response = requests.put(self.stage5_url, json=self.data, headers=self.headers)

                if response.status_code == 200:
                    #self.Generate_UID()
                    self.processed_params += 1

                    print("Data sent successfully")
                    print(response.json())  # prints the response content in JSON format


                    self.sent_data = True  # Flag that data has been sent successfully
                else:
                    print(f"Failed to send data: {response.status_code}")
                    print(response.text)  # prints the response body (error message)
            except requests.exceptions.RequestException as e:  # Catch any requests-related errors
                print("Error:", e)  # prints the error details
        else:
            if self.sent_data:
                print("Data has already been sent.")

    def Generate_UID(self):
        url = "http://192.168.2.253:6101/api/generate_uid"

        # Define the data to be sent in the request body
        data = {
            "IMEI": self.data['IMEI'],
            "ICCID": self.data['ICCID'],
            "model_name": self.data['model_name'],
            "QR_code": self.data['QR_code']
        }

        # Send the POST request
        response = requests.post(url, json=data)

        # Check the response status and content
        if response.status_code == 200:
            print("Request was successful.")

            # Assuming the response is in JSON format
            response_json = response.json()  # Get the JSON response as a dictionary

            print("Response:", response_json)

            # Now safely fetch the 'UIN' from the response
            UIN = response_json.get('UIN')
            self.data['UIN'] = UIN
            # Check if the UID was successfully assigned
            print("UID:", self.data['UID'])

            if self.data['UIN']:
                self.collect_sticker_prntng_params()
                print("UID (after processing):", self.data['UIN'])
            else:
                print("UIN not found in the response.")
        else:
            print("Request failed with status code:", response.status_code)
            print("Response:", response.text)

    def collect_sticker_prntng_params(self):
        stkr_prntg_var = [self.barcode,self.data.get('IMEI'),self.data.get('ICCID'),self.data.get('UIN')]
        print(stkr_prntg_var)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon("AEPL_Logo.png"))

    # Initialize UI
    ui = Ui_FinalTestingUtility()
    main_window = QMainWindow()
    ui.setupUi(main_window)

    # Initialize CAN_Data and pass the ui to it
    can_data_obj = CAN_Data(ui=ui)  # Only pass ui to CAN_Data

    # Initialize Worker without ui
    worker_obj = Worker(can_data=can_data_obj)  # Worker does not need ui

    
    # Initialize MyClass
    processor = MyClass(can_data_obj,worker_obj)
    processor.show()

    # Start console input listener
    # console_thread = threading.Thread(target=worker_obj.initialize_can_bus, daemon=True)
    # console_thread.start()
    app.exec()
    #sys.exit()

