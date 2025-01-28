
import can
import time
import binascii
import sys
import os
import requests
import datetime
from datetime import datetime, timezone
from datetime import timedelta
import pytz
import resources_rc
import threading
from PyQt5.QtCore import Qt
from PyQt5.QtGui  import QCursor
from PyQt5.QtCore import QPoint
from PyQt5.QtCore import QObject, pyqtSignal, QThread
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta,timezone
from PyQt5.QtWidgets import QMainWindow, QApplication,QMessageBox
from PyQt5.QtGui import QTextCursor, QTextBlockFormat
from PyQt5.QtCore import QDateTime
from PyQt5.QtCore import QTimer


# Expected CAN IDs and their frame counts
expected_frame_counts = {0x100: 3, 0x101 :3, 0x103 : 4, 0x104 : 4,0x105 :3, 0x106 :3 , 0x115 : 1, 0x116 : 1,0x109 :1,0x110:1, 0x112 :2,0x113 :1,0x114:5
                         ,0x102 : 1,0x119 :1, 0x121 : 1 , 0x122 : 1, 0x123:1}

# Initialize received_frames with empty lists for each CAN ID
received_frames = {0x100: [],0x101 : [] , 0x103 :[], 0x104 : [] ,0x105 :[],0x106 :[] , 0x115 :[], 0x116 : [],0x109:[],0x110:[],0x112:[],0x113:[],0x114:[]
                   ,0x102:[] , 0x119 :[] , 0x121 :[] ,0x122 :[] ,0x123 : []}



class Worker(QObject):

   
    update_103_singal = pyqtSignal(str)
    update_104_singal = pyqtSignal(str)
    update_106_singal = pyqtSignal(str)
    update_105_singal = pyqtSignal(str)
    update_101_singal = pyqtSignal(str)
    update_100_singal = pyqtSignal(str)
    update_110_singal = pyqtSignal(int,int,int)
    update_112_singal = pyqtSignal(str)
    update_109_singal = pyqtSignal(int,int)
    update_115_singal = pyqtSignal(float)
    update_116_singal = pyqtSignal(str)
    update_113_singal = pyqtSignal(int,int)
    update_114_singal = pyqtSignal(str,str,str,str)
    update_102_singal = pyqtSignal(timedelta)
    update_121_singal = pyqtSignal(bool,bool,bool,bool,int,int,int,int)
    update_122_signal = pyqtSignal(str)
    update_123_singal = pyqtSignal(bool,bool,int)
    update_119_singal = pyqtSignal(int,int)
    update_retryUI_signal = pyqtSignal(str,bool)
    
    def __init__(self,can_data=None):
        super().__init__()
        self.Can_data =can_data
        self.bus =None
        self.busy = False
        
        #self.ui.pushButton_8.clicked.connect(self.start_functions)
        # self.ui.pushButton_2.clicked.connect(self.save_to_excel)
        # self.ui.pushButton_9.clicked.connect(self.failed_func)
        #self.initialize_can_bus()
        self.current_datetime =None
        self.qc_head = None
        self.operator = None
        self.IMEI_ascii= None
        self.ICCID_ascii = None
        self.appln_ver = None
        self.BL_ver =None
        self.GSM_ver = None
        self.Gps_ver = None
        self.Int_vtg = None
        self.cleaned_IntBatVtg =None
        self.mains_vtg = None
        self.mains_vtg_str = None
        self.Gps_status = None
        self.No_of_Sat = None
        self.CREG = None
        self.CGREG = None
        self.CSQ = None
        self.operatorName = None
        self.MQTT_status = None
        self.No_of_LogInPacket =None
        self.chip_No = None
        self.frame1 = None
        self.frame2 = None
        self.frame3 = None
        self.RTC = None
        self.time_difference = None
        self.EpochToCurrentTime =None
        self.tamper = None
        self.mains_vtg_float = None

        self.IntVtg_result = None
        self.Gps_result = None
        self.RTC_result =None
        self.GSM_result = None
        self.DIs_result = None
        self.IGN_result = None
        self.Tamper_result = None
        self.MEMS_result = None
        self.MQTT_result = None
        self.overall_result=None
        self.MQTT_result=None
        self.WDT_result = None

        self.DI1_H =None
        self.DI1_L= None
        self.IGN = None
        self.DI1_status = False
        self.DI2_status = False
        self.DI3_status = False
        self.DI1_seen_0 = False
        self.DI1_seen_1 = False
        self.DI2_seen_0 = False
        self.DI2_seen_1 = False
        self.DI3_seen_0 = False
        self.DI3_seen_1 = False
        self.CREG_found = False
        self.CGREG_found = False
        self.CSQ_found = False
        self.operator_found = False
        self.IGN_seen_0 = False
        self.IGN_seen_1 = False
        self.No_of_Sat2 = None
        self.concatenated_hex = None
        self.concatenated_satellites_decimal = None
        self.elapsed_time = 0
        #self.retry=0
        self.failFunc_list =[]
        self.Flash_result =None
        self.device_id = None
        self.device_id_found = False
        self.device_id_true = False
        self.erase_status = None
        self.erase_status_found = False
        self.erase_status_true = False
        self.read_status = None
        self.read_status_found = False
        self.read_status_true = False
        self.write_status = None
        self.write_status_found = False
        self.write_status_true = False
        self.watchdog_reboot = None
        self.watchdogreboot_flag = False
        self.watchdog_reboot_count =None
        self.watchdog_reboot_count_dec = None
        self.watchdogrebootCount_flag = False
        self.fun123_checkedin= 0
        self.retry = 0
        self.reboot_str ="Please wait device is rebooting and requesting a boot count"
        self.retry_str = "Starting retries..."
        self.max_retries_flag= False
        self.WDT_MSG_flag = False
        
        # Initialize flags
        self.function100_done = False
        self.function101_done = False
        self.function103_done = False
        self.function104_done = False
        self.function105_done = False
        self.function106_done = False
        self.function115_done = False
        self.function116_done = False
        self.function109_done = False
        self.function110_done = False
        self.function112_done = False
        self.function113_done = False
        self.function114_done = False
        self.function102_done = False
        self.DIs_func_done  =   False
        self.function121_done = False
        self.function122_done = False
        self.function123_done = False

        # Timer for delays
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)  # Ensure it only fires once
        self.timer.timeout.connect(self.execute_next_function)


    # def initialize_can_bus(self):
    #     try:
    #         # Initialize the bus once, not inside each function
    #         self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
    #         #print(f"CAN Bus initialized: {self.bus.channel_info}")
    #     except can.CanError as e:
    #         print(f"Error initializing CAN bus: {str(e)}")
    #         self.bus = None  # Set bus to None if there's an initialization error

    #     self.start_functions()

    def start_functions(self):
    
        """Start the series of functions when the button is clicked."""
            # Reset flags
        self.function100_done = False
        self.function101_done = False
        self.function103_done = False
        self.function104_done = False
        self.function105_done = False
        self.function106_done = False
        self.function115_done = False
        self.function116_done = False
        self.function109_done = False
        self.function110_done = False
        self.function112_done = False
        self.function113_done = False
        self.function114_done = False
        self.function102_done = False
        self.DIs_func_done =  False
        self.function121_done = False
        self.function122_done = False
        self.function123_done = False
                
        # Call the first function
        self.fun_0x103()
        

    def fun_0x100(self,retry_mode =False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
            
        self.busy = True  # Mark the system as busy

        try:
            msg = can.Message(arbitration_id=0x100, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            
            # Wait for the response
            for i in range(expected_frame_counts[0x100]):
                message = self.bus.recv(timeout=2)  # 2 second timeout for each frame
                if message:
                    received_frames[0x100].append(message)
                
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x100]) == expected_frame_counts[0x100]:
                frames = received_frames[0x100]
                frames.sort(key=lambda x: x.data[0])  
                complete_message = b''.join(frame.data[1:] for frame in frames)
                IMEI = complete_message[:15]

                try:
                  self.IMEI_ascii = IMEI.decode('ascii')  
                  print(f"Extracted IMEI (ASCII): {self.IMEI_ascii}")
                  self.update_100_singal.emit(self.IMEI_ascii) 
                
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x100. Expected {expected_frame_counts[0x100]}, but received {len(received_frames[0x100])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")

        finally:
            self.busy = False  # Mark the system as not busy
            self.function100_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x100].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()

       
        
    def fun_0x101(self,retry_mode =False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
            
        self.busy = True  # Mark the system as busy

        try:
            msg = can.Message(arbitration_id=0x101, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)        
            self.bus.send(msg)
            
            for i in range(expected_frame_counts[0x101]):
                message = self.bus.recv(timeout=2)  
                
                if message:
                    received_frames[0x101].append(message)
                    
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x101]) == expected_frame_counts[0x101]:
                frames = received_frames[0x101]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Reassembled message for CAN ID 0x101: {complete_message.hex()}")
 
                ICCID = complete_message[:20]
                #print(f"Extracted IMEI: {ICCID.hex()}")
 
                try:
                  self.ICCID_ascii = ICCID.decode('ascii')  # Decode bytes into ASCII string
                  print(f"Extracted ICCID (ASCII): {self.ICCID_ascii}")
                  
                  self.update_101_singal.emit(self.ICCID_ascii)
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x101. Expected {expected_frame_counts[0x101]}, but received {len(received_frames[0x101])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function101_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x101].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
           


    def fun_0x103(self, retry_mode = False):

        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
 
        self.busy = True  # Mark the system as busy
        
        try:
            msg = can.Message(arbitration_id=0x103, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x103]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x103].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x103. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x103]) == expected_frame_counts[0x103]:
                frames = received_frames[0x103]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Reassembled message for CAN ID 0x100: {complete_message.hex()}")
                
                try:
                  self.appln_ver = complete_message.decode('ascii').rstrip('\x00').strip() # Decode bytes into ASCII string
                  #print('appln ver ASCII :',self.appln_ver)
                  print(f"Application version: {repr(self.appln_ver)}")
                  self.update_103_singal.emit(self.appln_ver)

                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x103. Expected {expected_frame_counts[0x103]}, but received {len(received_frames[0x103])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function103_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x103].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
          
       

    def fun_0x104(self,retry_mode = False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
    
        self.busy = True  # Mark the system as busy

        try:
            msg = can.Message(arbitration_id=0x104, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x104]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x104].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x104. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x104]) == expected_frame_counts[0x104]:
                frames = received_frames[0x104]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Reassembled message for CAN ID 0x100: {complete_message.hex()}")
 
                try:
                  self.BL_ver = complete_message.decode('ascii').rstrip('\x00').strip() # Decode bytes into ASCII string
                  #print('appln ver ASCII :',self.appln_ver)
                  print(f"Bootloader version: {repr(self.BL_ver)}")
                  self.update_104_singal.emit(self.BL_ver)
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x104. Expected {expected_frame_counts[0x104]}, but received {len(received_frames[0x104])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function104_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x104].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
            

    def fun_0x105(self , retry_mode =False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
 
        self.busy = True  # Mark the system as busy

        try:
            msg = can.Message(arbitration_id=0x105, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)  
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x105]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x105].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x105]) == expected_frame_counts[0x105]:
                frames = received_frames[0x105]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Complete for CAN ID 0x105: {complete_message.hex()}")
 
                try:
                  self.Gps_ver = complete_message.decode('ascii').rstrip('\x00').strip()  # Decode bytes into ASCII string
                  print('GPS ver ASCII :',self.Gps_ver)
                  self.update_105_singal.emit(self.Gps_ver)
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x105. Expected {expected_frame_counts[0x105]}, but received {len(received_frames[0x105])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function105_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x105].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
           


    def fun_0x106(self , retry_mode = False):
        print('fun 0x106 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
 
        self.busy = True  # Mark the system as busy

        try:
            msg = can.Message(arbitration_id=0x106, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x106]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                   
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x106].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x106]) == expected_frame_counts[0x106]:
                frames = received_frames[0x106]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Complete message for CAN ID 0x106: {complete_message.hex()}")
 
                try:
                  self.GSM_ver = complete_message.decode('ascii').rstrip('\x00').strip()
                  print('GSM ver ASCII :',{repr(self.GSM_ver)})
                  self.update_106_singal.emit(self.GSM_ver)
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x106. Expected {expected_frame_counts[0x106]}, but received {len(received_frames[0x106])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function106_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x106].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
            


    def fun_0x115(self , retry_mode = False):
        print('115 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
        
        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x115, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response
            print('hex value of mains :', message)

            if message:
               
                exttracted_mains = message.data[1:]
                # Process the received message
                #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
            
                # Decode the received message and update the UI
                self.mains_vtg = exttracted_mains.decode('ascii')  # Decode bytes into ASCII string
                print('Mains vtg:', {repr(self.mains_vtg)})
                cleaned_mains_vtg = self.mains_vtg.replace('\x00', '').strip()
                self.mains_vtg_float = float(cleaned_mains_vtg)  # Convert to float
                
                self.update_115_singal.emit(self.mains_vtg_float)

           
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x115. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function115_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x115].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
            

    def fun_0x116(self, retry_mode = False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x116, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                
                exttracted_IntVtg = message.data[1:]
                # Process the received message
                #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
            
                # Decode the received message and update the UI
                self.Int_vtg = exttracted_IntVtg.decode('ascii')  # Decode bytes into ASCII string
                self.cleaned_IntBatVtg = self.Int_vtg .replace('\x00', '').strip()
                print('Internal vtg:', self.Int_vtg)

                self.update_116_singal.emit(self.cleaned_IntBatVtg)
                
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x116. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function116_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x116].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()

    def fun_0x109(self,retry_mode=False):
       
        print('109 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x109, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                
                # Get GPS status and display it
                self.Gps_status = message.data[1]
                print('Gps status :', self.Gps_status)
                
                # Get the two hex values for the number of satellites
                self.No_of_Sat = hex(message.data[2])
                print('1st byte of no. of sat', self.No_of_Sat)
                self.No_of_Sat2 = hex(message.data[3])
                print('2nd byte of no. of sat', self.No_of_Sat2)
    
                # Concatenate the two hex values (removing the '0x' part before concatenation)
                self.concatenated_hex = self.No_of_Sat[2:] + self.No_of_Sat2[2:]
                print('Concatenated hex value:', self.concatenated_hex)
    
                # Convert the concatenated hex string to a decimal value
                self.concatenated_satellites_decimal = int(self.concatenated_hex, 16)
                print('No. of sat (Decimal):', self.concatenated_satellites_decimal)

                self.update_109_singal.emit(self.Gps_status,self.concatenated_satellites_decimal)
               
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function109_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x109].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
          

    def fun_0x110(self,retry_mode = False):
        print('inside 110')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
            
        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x110, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                self.CREG = message.data[1]
                print('CREG :',self.CREG)
            
                self.CGREG = message.data[2]
                print('CGREG:',self.CGREG)
                
                self.CSQ =message.data[3]
                print('CSQ',self.CSQ)

                print(self.CREG_found, self.CGREG_found, self.CSQ_found)
                
                self.update_110_singal.emit(self.CREG,self.CGREG,self.CSQ)
            
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function110_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x110].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
           
    def fun_0x112(self , retry_mode = False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
 
        self.busy = True  # Mark the system as busy

        try:
            msg = can.Message(arbitration_id=0x112, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x112]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    #print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x112].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x112]) == expected_frame_counts[0x112]:
                frames = received_frames[0x112]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                #print(f"Complete for CAN ID 0x112: {complete_message.hex()}")
 
                try:
                  self.operatorName = complete_message.decode('ascii').rstrip('\x00').strip()  # Decode bytes into ASCII string
                  print('Operator Name :',{repr(self.operatorName)})
                  self.update_112_singal.emit(self.operatorName)
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x112. Expected {expected_frame_counts[0x112]}, but received {len(received_frames[0x112])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function112_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x112].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
           
    def fun_0x113(self , retry_mode =False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x113, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                self.MQTT_status = message.data[1]
                print('MQTT status :',self.MQTT_status)

                self.No_of_LogInPacket = message.data[2]
                print('No. of Login Packet:',self.No_of_LogInPacket)
                
                self.update_113_singal.emit(self.MQTT_status,self.No_of_LogInPacket)
                # self.MQTT_result = self.ui.plainTextEdit_21.toPlainText()
                # print('mqtt result',self.MQTT_result)
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

            #self.MQTT_result = self.ui.plainTextEdit_21.toPlainText()

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function113_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x113].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
           

    def fun_0x114(self, retry_mode=False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy
        try:
            # Create and send the message
            msg = can.Message(arbitration_id=0x114, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
            self.bus.send(msg)

            # Reset frames before starting to receive
            self.chip_No = None
            self.frame1 = None
            self.frame2 = None
            self.frame3 = None

            # Create a list to keep track of frames received so far
            #received_frames = 0

            # Wait for the response
            all_frames = []  # Initialize an empty list to collect all frames

            # Try receiving all frames
            for i in range(expected_frame_counts[0x114]):  # Ensure it defaults to 0 if the key is missing
                message = self.bus.recv(timeout=2)  # 2-second timeout for each frame
                print('Received frames', message)
                if message:
                    # Extract the frame, skipping the 0th byte
                    frame = message.data[1:].hex()  # Skip the 0th byte and convert the rest to hex
                    all_frames.append(frame)  # Append the frame to the list

            # Join all frames into a single string and print
            frames = ''.join(all_frames)
            print("All Frames:", frames)

            # Convert the concatenated hex frames into bytes
            byte_data = bytes.fromhex(frames)

            # Decode the byte data into an ASCII string
            ascii_string = byte_data.decode('ascii', errors='ignore')
            print("ASCII String:", ascii_string)

            # Split the ASCII string based on commas
            parts = ascii_string.split(",")

            # Separate the first part (e.g., "6C") into a new variable
            self.chip_No = parts[0] if len(parts) > 0 else ""

            # Assign the remaining parts to frames based on the number of commas
            self.frame1 = parts[1] if len(parts) > 1 else ""
            self.frame2 = parts[2] if len(parts) > 2 else ""
            self.frame3 = ",".join(parts[3:]) if len(parts) > 3 else ""

            # Print the frames for verification
            print("Chip No.",self.chip_No)
            print("Frame 1:", self.frame1)
            print("Frame 2:", self.frame2)
            print("Frame 3:", self.frame3)

            self.update_114_singal.emit(self.frame1,self.frame2,self.frame3 ,self.chip_No)

            #self.MEMS_result = self.ui.plainTextEdit_25.toPlainText()

        except Exception as e:
            # Handle any exceptions that occur during the process
            print(f"An error occurred: {e}")

        finally:
            self.busy = False  # Mark the system as not busy
            self.function114_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x114].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()


    def fun_0x121(self, retry_mode = False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x121, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            #print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                #print('flash msg :',message)

                self.device_id = message.data[1]
                print('Device id :',self.device_id)
                if self.device_id !=20:
                    self.device_id_found = False
                else:
                    self.device_id_found = True
                    self.device_id_true = True

                self.erase_status = message.data[2]
                print('Flash erase status:',self.erase_status)
                if self.erase_status != 1:
                    self.erase_status_found =False
                else:
                    self.erase_status_found = True
                    self.erase_status_true = True
                

                self.read_status =message.data[3]
                print('Flash read status',self.read_status)
                if self.read_status != 1:
                    self.read_status_found = False
                else:
                   self.read_status_found = True
                   self.read_status_true =True


                self.write_status =message.data[4]
                print('Flash write status',self.write_status)
                if self.write_status != 1:
                    self.write_status_found = False
                else:
                   self.write_status_found = True
                   self.write_status_true = True

                print(self.device_id, self.erase_status, self.read_status,self.write_status)
                
                self.update_121_singal.emit(self.device_id_true,self.erase_status_true,self.read_status_true,self.write_status_true,
                                            self.device_id,self.erase_status,self.read_status,self.write_status)
                
                #self.Flash_result = self.ui.plainTextEdit_51.toPlainText()
            
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            self.function121_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x121].clear()
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()
           


    def fun_0x102(self, retry_mode=False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy
        current_time_utc = datetime.now(pytz.utc)  # Get current time in UTC (offset-aware)
        current_time_ist = current_time_utc.astimezone(pytz.timezone('Asia/Kolkata'))  # Convert to IST
        print('Current time (IST):', current_time_ist)

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x102, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                print('RTC msg',message)
                RTC_data = message.data[1:5]
            
                self.RTC = int.from_bytes(RTC_data, byteorder='big')
                print('INT RTC:', self.RTC)

                # Convert the epoch time (RTC) to datetime in UTC
                epoch_time_utc = datetime.fromtimestamp(self.RTC, tz=pytz.utc)
                epoch_time_ist = epoch_time_utc.astimezone(pytz.timezone('Asia/Kolkata'))  # Convert to IST
                print('Epoch to current time (IST):', epoch_time_ist)

                # Compare the two times (both are now offset-aware)
                self.time_difference = current_time_ist - epoch_time_ist

                self.update_102_singal.emit(self.time_difference)
                
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x102. No response received.")

            #self.RTC_result = self.ui.plainTextEdit_13.toPlainText()

            # if self.Mains_result == 'Pass' and self.IntVtg_result == 'Pass' and self.Gps_result == 'Pass' and self.GSM_result == 'Pass' and self.RTC_result == 'Pass' and self.MEMS_result == 'Pass' and self.MQTT_result == 'Pass':
            #     self.ui.textEdit.setPlainText("Pass")
            #     self.ui.textEdit.setAlignment(Qt.AlignCenter)
            #     self.ui.textEdit.setStyleSheet("""Font-size:22px; font-weight: Bold; background-color: green""")
               
            # else:
            #     self.ui.textEdit.setPlainText("Fail")
            #     self.ui.textEdit.setAlignment(Qt.AlignCenter)
            #     self.ui.textEdit.setStyleSheet("""Font-size:22px; font-weight: Bold; background-color: red """)
        
            #self.overall_result = self.ui.textEdit.toPlainText()

            
        except can.CanError as e:
            print(f"CAN error: {str(e)}")

        finally:
            self.busy = False  # Mark the system as not busy
            self.function102_done = True
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            received_frames[0x102].clear()  # Clear any frames in the buffer for ID 0x102
            if not retry_mode:  # Skip calling execute_next_function during retries
                self.execute_next_function()

    def fun_0x122(self,retry_mode =False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy

        try:
            
            # Create the CAN message
            msg = can.Message(arbitration_id=0x122, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
               
                self.watchdog_reboot = message.data[1]
                print('Watchdog reboot status:', self.watchdog_reboot)

                if self.watchdog_reboot != 1:
                    self.watchdogreboot_flag = False
                    
                else:
                    self.watchdogreboot_flag = True
                    self.update_122_signal.emit(self.reboot_str)
                    # self.retry_timer = QTimer(self)
                    # self.retry_timer.timeout.connect(self.update_rebootStatus)
                    # self.retry_timer.setSingleShot(True)
                    # self.retry_timer.start(1000)
                    
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x102. No response received.")


        except can.CanError as e:
            print(f"CAN error: {str(e)}")

        finally:
            self.busy = False  # Mark the system as not busy
            if self.bus:  # Ensure we properly shut down the bus
                self.bus.shutdown()  # Properly shut down the bus
                self.bus = None  # Set bus to None
            print('inside finally block')
            received_frames[0x122].clear()  # Clear any frames in the buffer for ID 0x102
            self.function122_done = True
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x123)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(35000)
            #time.sleep(2)  # Sleep to allow processing
            
            #self.execute_next_function()  # Move on to the next function

    # def update_rebootStatus(self):
    #     self.ui.plainTextEdit_12.setPlainText(" Please wait device is rebooting...")
    #     self.ui.plainTextEdit_12.setStyleSheet("""
    #         font-size: 16px; 
    #         font-weight: bold; 
    #         color: red;
    #     """)

    
    def fun_0x123(self,retry_mode=False):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus is not initialized
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return
            
        self.busy = True  # Mark the system as busy

        # Initialize previous watchdog count if it doesn't exist
        if not hasattr(self, 'prev_watchdog_reboot_count_dec'):
            self.prev_watchdog_reboot_count_dec = None
        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x123, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
            
            # Send the message once
            self.bus.send(msg)
            print('123 send msg',msg)

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                print('received 123 hex msg',message)
                # Update the current watchdog reboot count with new data
                self.watchdog_reboot_count = message.data[1:5]
                self.watchdog_reboot_count_dec = int.from_bytes(self.watchdog_reboot_count, byteorder='big')
                print('Current watchdog reboot count decimal:', self.watchdog_reboot_count_dec)
                # Compare with previous count
                if self.prev_watchdog_reboot_count_dec is not None:
                    print('Previous watchdog reboot count decimal:', self.prev_watchdog_reboot_count_dec)
                    if self.watchdog_reboot_count_dec > self.prev_watchdog_reboot_count_dec:
                        self.watchdogrebootCount_flag = True
                        print("Watchdog reboot count is incremental as expected.")
                    elif self.watchdog_reboot_count_dec == self.prev_watchdog_reboot_count_dec:
                        self.watchdogrebootCount_flag = False
                        print("Error: Watchdog reboot count is the same as the previous count.")
                    else:
                        print("Error: Watchdog reboot count has decreased, which is unexpected!")
                else:
                    print("No previous watchdog reboot count available for comparison.")
                    # print("\n")
                    # print("########## Device reboot Starting #########")
                    # print(" Please wait device is rebooting...")

                # Update the previous watchdog reboot count
                self.prev_watchdog_reboot_count_dec = self.watchdog_reboot_count_dec
                self.fun123_checkedin +=1
                if self.fun123_checkedin ==2:
                    self.fun123_checkedin =0

                    self.update_123_singal.emit(self.watchdogreboot_flag ,self.watchdogrebootCount_flag,self.watchdog_reboot_count_dec)

                    

                #     if self.Mains_result == 'Pass' and self.IntVtg_result == 'Pass' and self.Gps_result == 'Pass' and self.GSM_result == 'Pass' and self.RTC_result == 'Pass' and self.MEMS_result == 'Pass' and self.MQTT_result == 'Pass'and self.WDT_result == 'Pass':
                #         self.ui.textEdit.setPlainText("Pass")
                #         self.ui.textEdit.setStyleSheet("""Font-size:20px; font-weight: Bold; background-color: green""")
                #     else:
                #         self.ui.textEdit.setPlainText("Fail")
                #         self.ui.textEdit.setStyleSheet("""Font-size:20px; font-weight: Bold; background-color: red""")
                #     self.overall_result = self.ui.textEdit.toPlainText()
                else:
                    pass

            else:
                
                self.update_123_singal.emit(self.watchdogreboot_flag ,self.watchdogrebootCount_flag,self.watchdog_reboot_count_dec)



        finally:
                self.busy = False  # Mark the system as not busy
                self.function123_done = True
                if self.bus:  # Ensure we properly shut down the bus
                    self.bus.shutdown()  # Properly shut down the bus
                    self.bus = None  # Set bus to None
                received_frames[0x123].clear()  # Clear any frames in the buffer for ID 0x102
                if not retry_mode:  # Skip calling execute_next_function during retries
                    self.execute_next_function()



    # def DIs_func(self):
    #         if self.busy:  # Check if the system is busy
    #             print("System is busy, please wait...")
    #             return

    #         if self.bus is None:  # Check if the bus was initialized properly
    #             print("CAN Bus not initialized. Cannot send message.")
    #             return

    #         self.busy = True  # Mark the system as busy

    #         try:
    #             # Create the CAN message
    #             msg = can.Message(arbitration_id=0x119, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

    #             # Send the message once
    #             self.bus.send(msg)

    #             # Wait for a response with a timeout (e.g., 2 seconds)
    #             message = self.bus.recv(timeout=2)  # 2 seconds timeout for response
        
    #             if message:
    #                 self.IGN = message.data[1]
    #                 print('IGN :', self.IGN)
            
    #                 self.tamper = message.data[2]
    #                 self.ui.Tamp_L.setPlainText(str(self.tamper))
    #                 print('Tamper:', self.tamper)
            
    #                 self.DI1 = message.data[3]
    #                 print('DI1 :', self.DI1)
            
    #                 self.DI2 = message.data[4]
    #                 print('DI2 :', self.DI2)

    #                 self.DI3 = message.data[5]
    #                 print('DI3 :', self.DI3)
            
    #             else:
    #                 # If no message is received within the timeout period
    #                 print(f"Timeout waiting for message for CAN ID 0x119. No response received.")

    #             if self.tamper != 0:
    #                 self.ui.Tamp_L.setStyleSheet("background-color : red")
    #                 self.ui.plainTextEdit_47.setPlainText("Fail")
    #                 self.ui.plainTextEdit_47.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
    #             else:
    #                 self.ui.Tamp_L.setStyleSheet("background-color : white")
    #                 self.ui.plainTextEdit_47.setPlainText("Pass")
    #                 self.ui.plainTextEdit_47.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")

    #             self.Tamper_result = self.ui.plainTextEdit_47.toPlainText()

    #             # Update UI fields if they are empty
    #             if not self.ui.DI1_H_3.toPlainText():
    #                 self.ui.DI1_H_3.setPlainText(str(self.IGN))
                    
    #             else:
    #                 self.ui.IGN_H.setPlainText(str(self.IGN))

    #             if not self.ui.DI1_H_6.toPlainText():
    #                 self.ui.DI1_H_6.setPlainText(str(self.DI1))
    #             else:
    #                 self.ui.DI1_H_7.setPlainText(str(self.DI1))

    #             if not self.ui.DI1_H_4.toPlainText():
    #                 self.ui.DI1_H_4.setPlainText(str(self.DI2))
    #             else:
    #                 self.ui.DI1_H_5.setPlainText(str(self.DI2))

    #             if not self.ui.DI1_H_8.toPlainText():
    #                 self.ui.DI1_H_8.setPlainText(str(self.DI3))
    #             else:
    #                 self.ui.DI_H.setPlainText(str(self.DI3))

    #             self.ui.plainTextEdit_12.appendPlainText(f"IGN: {str(self.IGN)}\n")
    #             self.ui.plainTextEdit_12.appendPlainText(f"Tamper: {str(self.tamper)}\n")
    #             self.ui.plainTextEdit_12.appendPlainText(f"DI1,DI2,DI3: {self.DI1}, {self.DI2}, {self.DI3}")

    #             # Track whether we have seen both states (0 and 1) for each DI
    #             # Check and update DI1 status
    #             if self.DI1 == 0:
    #                 self.DI1_seen_0 = True  # Mark that we have seen 0 for DI1
    #             elif self.DI1 == 1:
    #                 self.DI1_seen_1 = True  # Mark that we have seen 1 for DI1

    #             # Check and update DI2 status
    #             if self.DI2 == 0:
    #                 self.DI2_seen_0 = True  # Mark that we have seen 0 for DI2
    #             elif self.DI2 == 1:
    #                 self.DI2_seen_1 = True  # Mark that we have seen 1 for DI2

    #             # Check and update DI3 status
    #             if self.DI3 == 0:
    #                 self.DI3_seen_0 = True  # Mark that we have seen 0 for DI3
    #             elif self.DI3 == 1:
    #                 self.DI3_seen_1 = True  # Mark that we have seen 1 for DI3

    #             if self.IGN == 0:
    #                 self.IGN_seen_0 =True
    #             else:
    #                 self.IGN_seen_1 = True

    #             # Use QTimer to periodically check if both 0 and 1 have been seen for each DI
    #             self.timer = QTimer(self)
    #             self.timer.setInterval(1000)  # 1000 ms = 1 second
    #             self.timer.timeout.connect(self.check_flags)  # Connect timeout to the check_flags function
    #             self.timer.start(1000)  # Check every second (1000 ms)

    #         except can.CanError as e:
    #             print(f"CAN error: {str(e)}")
    
    #         finally:
    #             self.busy = False  # Mark the system as not busy
    #             received_frames[0x119].clear()
    #             self.DIs_func_done = True
    #             time.sleep(2)


    # def check_flags(self):
    #     # This method will be called every second
    #     #print(f"Checking flags: DI1_seen_0={self.DI1_seen_0}, DI1_seen_1={self.DI1_seen_1}, DI2_seen_0={self.DI2_seen_0}, DI2_seen_1={self.DI2_seen_1}, DI3_seen_0={self.DI3_seen_0}, DI3_seen_1={self.DI3_seen_1}")

    #     # Now check if all flags are True
    #     if self.DI1_status and self.DI2_status and self.DI3_status:
    #         self.timer.stop()  # Stop the timer when all flags are True
    #         print('timer stopped')
        
    #         # Now that all DI states are confirmed (both 0 and 1), determine the result
    #         if self.DI1_status and self.DI2_status and self.DI3_status:
    #             self.ui.plainTextEdit_22.setPlainText("Pass")
    #             self.ui.plainTextEdit_22.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
    #         else:
    #             self.ui.plainTextEdit_22.setPlainText("Fail")
    #             self.ui.plainTextEdit_22.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
        

    #     if self.IGN_seen_0 and self.IGN_seen_1:
    #         self.timer.stop()
    #         if self.IGN_seen_0 and self.IGN_seen_1:

    #             self.ui.plainTextEdit_26.setPlainText("Pass")
    #             self.ui.plainTextEdit_26.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: green""")
    #         else:
    #             self.ui.plainTextEdit_26.setPlainText("Fail")
    #             self.ui.plainTextEdit_26.setStyleSheet("""Font-size:16px; font-weight: Bold; background-color: red""")
    
    #     self.DIs_result = self.ui.plainTextEdit_22.toPlainText()
    #     self.IGN_result = self.ui.plainTextEdit_26.toPlainText()

    def DIs_func(self, retry_mode=False):
        if self.busy:  # Check if the system is busy
                print("System is busy, please wait...")
                return

        if self.bus is None:  # Check if the bus was initialized properly
            try:
                self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
                print("PCAN bus initialized successfully.")
            except can.CanError as e:
                print(f"Error initializing CAN bus: {str(e)}")
                return

        self.busy = True  # Mark the system as busy

        try:
                # Create the CAN message
                msg = can.Message(arbitration_id=0x119, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

                # Send the message once
                self.bus.send(msg)

                # Wait for a response with a timeout (e.g., 2 seconds)
                message = self.bus.recv(timeout=2)  # 2 seconds timeout for response
        
                if message:
                    self.IGN = message.data[1]
                    print('IGN :', self.IGN)

                    self.tamper = message.data[2]
                    print('Tamper :',self.tamper)

                    self.update_119_singal.emit(self.IGN , self.tamper)

        except can.CanError as e:
                print(f"CAN error: {str(e)}")
    
        finally:
                self.busy = False  # Mark the system as not busy
                self.DIs_func_done = True
                received_frames[0x119].clear()
                if self.bus:  # Ensure we properly shut down the bus
                    self.bus.shutdown()  # Properly shut down the bus
                    self.bus = None  # Set bus to None
                if not retry_mode:  # Skip calling execute_next_function during retries
                    self.execute_next_function()


    def execute_next_function(self):
        """Check which function is done and call the next one."""
        if self.function103_done and not self.function104_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x104)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function104_done and not self.function106_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x106)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function106_done and not self.function105_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x105)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)

        elif self.function105_done and not self.function101_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x101)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)

        elif self.function101_done and not self.function100_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x100)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
    
        elif self.function100_done and not self.function110_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x110)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
        
        elif self.function110_done and not self.function112_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x112)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
           
        elif self.function112_done and not self.function109_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x109)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
           
        elif self.function109_done and not self.function115_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x115)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function115_done and not self.function116_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x116)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            

        elif self.function116_done and not self.function113_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x113)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function113_done and not self.function114_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x114)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function114_done and not self.function102_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x102)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function102_done and not self.function121_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x121)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        elif self.function121_done and not self.DIs_func_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.DIs_func)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            

        elif self.DIs_func and not self.function123_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x123)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
           
        elif self.function123_done and not self.function122_done:
            self.retry_timer = QTimer(self)
            self.retry_timer.timeout.connect(self.fun_0x122)
            self.retry_timer.setSingleShot(True)
            self.retry_timer.start(1000)
            
        else:
            print("All functions completed.")
            # You can enable a button or perform other tasks once all functions are done

    def failed_func(self):
        print("Inside fail_func")
        self.retry = 0
        self.max_retries_flag = False
        self.update_retryUI_signal.emit(self.retry_str,self.max_retries_flag)
        # self.ui.plainTextEdit_12.setPlainText("Starting retries...")
        # self.ui.plainTextEdit_12.setStyleSheet("""
        #     font-size: 16px; 
        #     font-weight: bold; 
        #     color: red;
        # """)
        
        self.failFunc_list = []  # Ensure the list is empty before populating it
        self.fail_attempts = {}  # A dictionary to track the retry attempts for each function

        # Populate the list of failed functions and initialize their retry attempts
        if self.Mains_result == 'Fail':
            self.failFunc_list.append(self.fun_0x115)
            self.fail_attempts[self.fun_0x115] = 0
        if self.IntVtg_result == 'Fail':
            self.failFunc_list.append(self.fun_0x116)
            self.fail_attempts[self.fun_0x116] = 0
        if self.Gps_result == 'Fail':
            self.failFunc_list.append(self.fun_0x109)
            self.fail_attempts[self.fun_0x109] = 0
        if self.GSM_result == 'Fail':
            self.failFunc_list.append(self.fun_0x110)
            self.fail_attempts[self.fun_0x110] = 0
        # if self.IGN_result == 'Fail':
        #     self.failFunc_list.append(self.DIs_func)
        # if self.Tamper_result == 'Fail':
        #     self.failFunc_list.append(self.Dis_func)
        if self.Flash_result == 'Fail':
            self.failFunc_list.append(self.fun_0x121)
            self.fail_attempts[self.fun_0x121] = 0
        if self.MEMS_result == 'Fail':
            self.failFunc_list.append(self.fun_0x114)
            self.fail_attempts[self.fun_0x114] = 0
        if self.MQTT_result == 'Fail':
            self.failFunc_list.append(self.fun_0x113)
            self.fail_attempts[self.fun_0x113] = 0
        if self.RTC_result == 'Fail':
            self.failFunc_list.append(self.fun_0x102)
            self.fail_attempts[self.fun_0x102] = 0
        if self.WDT_result == 'Fail':
            self.failFunc_list.extend([self.fun_0x123])
            self.fail_attempts[self.fun_0x123] = 0

        # Print the failed function list
        print("Initial failed functions:", [func.__name__ for func in self.failFunc_list])

      

        # Create and start QTimer for retries
        self.retry_timer = QTimer(self)
        self.retry_timer.timeout.connect(self.retry_iteration)
        self.retry_timer.start(20000)  # Start with a 20-second interval

        # Trigger the first retry immediately
        #self.retry_iteration()

    def retry_iteration(self):
        if self.retry >= 3:
            print("Max retries reached. Stopping...")
            self.max_retries_flag= True
            #QMessageBox.warning(self, "Warning", "Max retries reached. Stopping retries.")
            self.update_retryUI_signal.emit(self.retry_str,self.max_retries_flag)
            self.retry_timer.stop()
            #self.ui.pushButton_9.setEnabled(True)
            return

        print(f"Starting retry iteration {self.retry + 1}...")
        

        # Iterate through the functions in the failure list
        for func in list(self.failFunc_list):  # Using list() to avoid modifying the list while iterating
            # Check if the function has passed already
            if self.is_flag_passed(func):
                print(f"Function {func.__name__} already passed. Skipping...")
                self.failFunc_list.remove(func)  # Remove passed functions from the retry list
                
                continue

            try:
                # Retry the function (pass `retry_mode=True` to prevent execution of next function)
                success = func(retry_mode=True)  # If successful, function should return True
                if success:
                    print(f"Function {func.__name__} passed.")
                    self.update_flag(func, 'Pass')  # Update the flag to "Pass"
                    self.failFunc_list.remove(func)  # Remove from the failure list
                   
                else:
                    print(f"Function {func.__name__} failed, will retry.")
            except Exception as e:
                print(f"Error while calling {func.__name__}: {e}")
                continue

        # If there are no functions left to retry, stop the retry timer
        if not self.failFunc_list:
            print("All functions succeeded. Stopping retries.")
            self.retry_timer.stop()
            #self.update_retryUI_signal.emit(self.retry_str,self.self.max_retries_flag)
            return

        self.retry += 1



    def is_flag_passed(self, func):
        if func == self.fun_0x115:
            return self.Mains_result == 'Pass'
        elif func == self.fun_0x116:
            return self.IntVtg_result == 'Pass'
        elif func == self.fun_0x109:
            return self.Gps_result == 'Pass'
        elif func == self.fun_0x110:
            return self.GSM_result == 'Pass'
        elif func == self.fun_0x121:
            return self.Flash_result == 'Pass'
        elif func == self.fun_0x114:
            return self.MEMS_result == 'Pass'
        elif func == self.fun_0x113:
            return self.MQTT_result == 'Pass'
        elif func == self.fun_0x102:
            return self.RTC_result == 'Pass'
        elif func == self.fun_0x114:
            return self.MEMS_result == 'Pass'
        elif func == self.fun_0x123 or self.fun_0x122:
            return self.WDT_result == 'Pass'
        
        
        return False

            

    # def clean_string(self,input_string):
    
    #         # Filter out characters with ASCII values from 0-31 and 127
    #         return ''.join(c for c in input_string if 31 < ord(c) < 127)

    # def save_to_excel(self):
    #     # Get the current date in YYYY-MM-DD format
    #     current_date = datetime.now().strftime("%Y-%m-%d")

    #     # Create a file path with the date appended to the filename
    #     current_directory = os.getcwd()
    #     file_path = os.path.join(current_directory, f"Final_Testing_DeviceStatus_{current_date}.xlsx")

    #     # If the file exists, load the existing workbook; otherwise, create a new one
    #     if os.path.exists(file_path):
    #         wb = load_workbook(file_path)
    #         ws = wb.active
    #     else:
    #         wb = Workbook()
    #         ws = wb.active

    #         # Set the headers for the columns (only if the file is being created for the first time)
    #         headers = ['Date', 'IMEI', 'ICCID', 'Application Version','BL version', 'GSM Version',
    #                'GPS Version', 'Mains vtg', 'Int_Bat vtg', 'GPS status', 'No.of Sat', 'CREG', 'CGREG', 'CSQ',
    #                'Operator', 'MQTT', 'No.Of Login packet', 'MEMS Xa', 'MEMS Ya', 'MEMS Za', 'Mains result', 'IntBat result'
    #                ,'Gps result', 'GSM result','IGN result','Tamper result','FlashMemory result','DIs result','DOs result',
    #                'AnalogVolt result','MEMS result','MQTT result','RTC result','WDT result','Overall Result']
    #         ws.append(headers)  # Append headers as the first row

    #     self.current_datetime =self.ui.operator_Input_2.toPlainText()

    

    # # Clean the data before inserting into the worksheet
    #     data = [
    #     self.clean_string(str(self.current_datetime)) if self.current_datetime is not None else 'Not found',
    #     self.clean_string(self.IMEI_ascii) if self.IMEI_ascii is not None else 'Not found',
    #     self.clean_string(self.ICCID_ascii) if self.ICCID_ascii is not None else 'Not found',
    #     self.clean_string(self.appln_ver) if self.appln_ver is not None else 'Not found',
    #     self.clean_string(self.BL_ver) if self.BL_ver is not None else 'Not found',
    #     self.clean_string(self.GSM_ver) if self.GSM_ver is not None else 'Not found ',
    #     self.clean_string(self.Gps_ver) if self.Gps_ver is not None else 'Not found',
    #     self.clean_string(self.mains_vtg) if self.mains_vtg is not None else 'Not found',
    #     self.clean_string(self.Int_vtg) if self.Int_vtg is not None else 'Not found',
    #     self.clean_string(str(self.Gps_status)) if self.Gps_status is not None else 'Not found',
    #     self.clean_string(str(self.concatenated_satellites_decimal)) if self.concatenated_satellites_decimal is not None else 'Not found',
    #     self.clean_string(str(self.CREG)) if self.CREG is not None else 'Not found',
    #     self.clean_string(str(self.CGREG)) if self.CGREG is not None else 'Not found',
    #     self.clean_string(str(self.CSQ)) if self.CSQ is not None else 'Not found',
    #     self.clean_string(str(self.operatorName)) if self.operatorName is not None else 'Not found',
    #     self.clean_string(str(self.MQTT_status)) if self.MQTT_status is not None else 'Not found',
    #     self.clean_string(str(self.No_of_LogInPacket)) if self.No_of_LogInPacket is not None else 'Not found',
    #     self.clean_string(str(self.frame1)) if self.frame1 is not None else 'Not found',
    #     self.clean_string(str(self.frame2)) if self.frame2 is not None else 'Not found',
    #     self.clean_string(str(self.frame3)) if self.frame3 is not None else 'Not found',
    #     self.clean_string(str(self.Mains_result)) if self.Mains_result is not None else 'Not found',
    #     self.clean_string(str(self.IntVtg_result)) if self.IntVtg_result is not None else 'Not found',
    #     self.clean_string(str(self.Gps_result)) if self.Gps_result is not None else 'Not found',
    #     self.clean_string(str(self.GSM_result)) if self.GSM_result is not None else 'Not found',
    #     self.clean_string(str(self.IGN_result)) if self.IGN_result is not None else 'Not found',
    #     self.clean_string(str(self.Tamper_result)) if self.Tamper_result is not None else 'Not found',
    #     self.clean_string(str(self.Flash_result)) if self.Flash_result is not None else 'Not found',
    #     self.clean_string(str(self.DIs_result)) if self.DIs_result is not None else 'Not found',
    #     self.clean_string(str(self.Dos_result)) if self.DOs_result is not None else 'Not found',
    #     self.clean_string(str(self.AnalogVolt_result)) if self.AnalogVolt_result is not None else 'Not found',
    #     self.clean_string(str(self.MEMS_result)) if self.MEMS_result is not None else 'Not found',
    #     self.clean_string(str(self.MQTT_result)) if self.MQTT_result is not None else 'Not found',
    #     self.clean_string(str(self.RTC_result)) if self.RTC_result is not None else 'Not found',
    #     self.clean_string(str(self.WDT_result)) if self.WDT_result is not None else 'Not found',
    #     self.clean_string(str(self.overall_result)) if self.overall_result is not None else 'Not found'
    #     ]

    #     # Append the data to the next available row
    #     ws.append(data)

    #     try:
    #         # Save the workbook to the specified path
    #         wb.save(file_path)
    #         print(f"Data saved to {file_path}")
    #         QMessageBox.information(self, "Success", f"Data successfully saved to {file_path}")

    #         # Clear the UI inputs after saving data
    #         self.clear_ui()

    #     except Exception as e:
    #         print(f"Error saving data to Excel: {str(e)}")
    #         QMessageBox.warning(self, "Error", f"Failed to save data to Excel: {str(e)}")


 

class CAN_Data(QMainWindow):
    def __init__(self, ui):
        super().__init__()
        self.ui = ui
        self.stackedWidget = self.ui.stackedWidget

        #self.worker = Worker()

        # self.Mains_result = None
        # self.IntVtg_result = None    
        # self.Gps_result = None
        # self.GSM_result = None
        # self.IGN_result = None
        # self.Tamper_result = None
        # self.Flash_result = None
        # self.MEMS_result = None
        # self.MQTT_result = None
        # self.RTC_result = None
        # self.WDT_result = None
        # self.overall_result =None

        
    def clear_ui(self):

        # self.ui.QC_input_2.clear()
        # self.ui.QC_input_2.setStyleSheet("background-color: white;")
        self.ui.barcode_Input_2.clear()
        self.ui.barcode_Input.setStyleSheet("background-color: white;")
        self.ui.server_Input.clear()
        self.ui.server_Input.setStyleSheet("background-color: white;")
        self.ui.barcode_Input.clear()
        self.ui.barcode_Input.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_5.clear()
        self.ui.plainTextEdit_5.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_6.clear()
        self.ui.plainTextEdit_6.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_8.clear()
        self.ui.plainTextEdit_8.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_9.clear()
        self.ui.plainTextEdit_9.setStyleSheet("background-color: white;")
        self.ui.Tamp_L.clear()
        self.ui.Tamp_L.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_10.clear()
        self.ui.plainTextEdit_10.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_11.clear()
        self.ui.plainTextEdit_11.setStyleSheet("background-color: white;")
        self.ui.CSQ.clear()
        self.ui.CSQ.setStyleSheet("background-color: white;")
        self.ui.CGREG.clear()
        self.ui.CGREG.setStyleSheet("background-color: white;")
        self.ui.CREG.clear()
        self.ui.CREG.setStyleSheet("background-color: white;")
        self.ui.Operator.clear()
        self.ui.Operator.setStyleSheet("background-color: white;")
        self.ui.NoOf_satellite.clear()
        self.ui.NoOf_satellite.setStyleSheet("background-color: white;")
        self.ui.Operator_2.clear()
        self.ui.Operator_2.setStyleSheet("background-color: white;")
        self.ui.MQTT.clear()
        self.ui.MQTT.setStyleSheet("background-color: white;")
        self.ui.mains_input_2.clear()
        self.ui.mains_input_2.setStyleSheet("background-color: white;")
        self.ui.IntBat_input_2.clear()
        self.ui.IntBat_input_2.setStyleSheet("background-color: white;")
        self.ui.Analog1_2.clear()
        self.ui.Analog1_2.setStyleSheet("background-color: white;")
        self.ui.MEMS_Xa.clear()
        self.ui.MEMS_Xa.setStyleSheet("background-color: white;")
        self.ui.MEMS_Ya.clear()
        self.ui.MEMS_Ya.setStyleSheet("background-color: white;")
        self.ui.MEMS_Za.clear()
        self.ui.MEMS_Za.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_12.clear()
        self.ui.plainTextEdit_12.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_23.clear()
        self.ui.plainTextEdit_23.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_29.clear()
        self.ui.plainTextEdit_29.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_28.clear()
        self.ui.plainTextEdit_28.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_21.clear()
        self.ui.plainTextEdit_21.setStyleSheet("background-color: white;")
        self.ui.DI1_H_3.clear()
        self.ui.DI1_H_3.setStyleSheet("background-color: white;")
        self.ui.IGN_H.clear()
        self.ui.IGN_H.setStyleSheet("background-color: white;")
        self.ui.DI1_H_6.clear()
        self.ui.DI1_H_6.setStyleSheet("background-color: white;")
        self.ui.DI1_H_7.clear()
        self.ui.DI1_H_7.setStyleSheet("background-color: white;")
        self.ui.DI1_H_4.clear()
        self.ui.DI1_H_4.setStyleSheet("background-color: white;")
        self.ui.DI1_H_5.clear()
        self.ui.DI1_H_5.setStyleSheet("background-color: white;")
        self.ui.DI1_H_8.clear()
        self.ui.DI1_H_8.setStyleSheet("background-color: white;")
        self.ui.DI_H.clear()
        self.ui.DI_H.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_24.clear()
        self.ui.plainTextEdit_24.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_22.clear()
        self.ui.plainTextEdit_22.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_26.clear()
        self.ui.plainTextEdit_26.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_47.clear()
        self.ui.plainTextEdit_47.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_25.clear()
        self.ui.plainTextEdit_25.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_31.clear()
        self.ui.plainTextEdit_31.setStyleSheet("background-color: white;")
        self.ui.textEdit.clear()
        self.ui.textEdit.setStyleSheet("background-color: white;")
        # self.ui.textEdit.clear()
        # self.ui.textEdit.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_13.clear()
        self.ui.plainTextEdit_13.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_51.clear()
        self.ui.plainTextEdit_51.setStyleSheet("background-color: white;")
        self.ui.plainTextEdit_14.clear()
        self.ui.plainTextEdit_14.setStyleSheet("background-color: white;")
        self.ui.MEMS_Xa_2.clear()
        self.ui.MEMS_Xa_2.setStyleSheet("background-color: white;")

  

    # def failed_func(self):
    #     print("Inside fail_func")
    #     # self.ui.plainTextEdit_12.setPlainText("Starting retries...")
    #     # self.ui.plainTextEdit_12.setStyleSheet("""
    #     #     font-size: 16px; 
    #     #     font-weight: bold; 
    #     #     color: red;
    #     # """)

    #     self.failFunc_list = []  # Ensure the list is empty before populating it
    #     self.fail_attempts = {}  # A dictionary to track the retry attempts for each function

    #     # Populate the list of failed functions and initialize their retry attempts
    #     if self.Mains_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x115)
    #         self.fail_attempts[self.worker.fun_0x115] = 0
    #     if self.IntVtg_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x116)
    #         self.fail_attempts[self.worker.fun_0x116] = 0
    #     if self.Gps_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x109)
    #         self.fail_attempts[self.worker.fun_0x109] = 0
    #     if self.GSM_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x110)
    #         self.fail_attempts[self.worker.fun_0x110] = 0
    #     # if self.IGN_result == 'Fail':
    #     #     self.failFunc_list.append(self.DIs_func)
    #     # if self.Tamper_result == 'Fail':
    #     #     self.failFunc_list.append(self.Dis_func)
    #     if self.Flash_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x121)
    #         self.fail_attempts[self.worker.fun_0x121] = 0
    #     if self.MEMS_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x114)
    #         self.fail_attempts[self.worker.fun_0x114] = 0
    #     if self.MQTT_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x113)
    #         self.fail_attempts[self.worker.fun_0x113] = 0
    #     if self.RTC_result == 'Fail':
    #         self.failFunc_list.append(self.worker.fun_0x102)
    #         self.fail_attempts[self.worker.fun_0x102] = 0
    #     if self.WDT_result == 'Fail':
    #         self.failFunc_list.extend([self.worker.fun_0x123, self.worker.fun_0x122])
    #         self.fail_attempts[self.worker.fun_0x123] = 0

    #     # Print the failed function list
    #     print("Initial failed functions:", [func.__name__ for func in self.failFunc_list])

    #     self.retry = 0  # Initialize retry count

    #     # Create and start QTimer for retries
    #     self.retry_timer = QTimer(self)
    #     self.retry_timer.timeout.connect(self.retry_iteration)
    #     self.retry_timer.start(20000)  # Start with a 20-second interval

    #     # Trigger the first retry immediately
    #     #self.retry_iteration()

    # def retry_iteration(self):
    #     if self.retry >= 3:
    #         print("Max retries reached. Stopping...")
    #         QMessageBox.warning(self, "Warning", "Max retries reached. Stopping retries.")
    #         self.retry_timer.stop()
    #         self.ui.pushButton_9.setEnabled(True)
    #         return

    #     print(f"Starting retry iteration {self.retry + 1}...")
    #     self.ui.pushButton_9.setEnabled(False)

    #     # Iterate through the functions in the failure list
    #     for func in list(self.failFunc_list):  # Using list() to avoid modifying the list while iterating
    #         # Check if the function has passed already
    #         if self.is_flag_passed(func):
    #             print(f"Function {func.__name__} already passed. Skipping...")
    #             self.failFunc_list.remove(func)  # Remove passed functions from the retry list
    #             self.update_ui_on_function_success(func)  # Ensure UI reflects success
    #             continue

    #         try:
    #             # Retry the function (pass `retry_mode=True` to prevent execution of next function)
    #             success = func(retry_mode=True)  # If successful, function should return True
    #             if success:
    #                 print(f"Function {func.__name__} passed.")
    #                 self.update_flag(func, 'Pass')  # Update the flag to "Pass"
    #                 self.failFunc_list.remove(func)  # Remove from the failure list
    #                 self.update_ui_on_function_success(func)  # Update the UI based on the passed function
    #             else:
    #                 print(f"Function {func.__name__} failed, will retry.")
    #         except Exception as e:
    #             print(f"Error while calling {func.__name__}: {e}")
    #             continue

    #     # If there are no functions left to retry, stop the retry timer
    #     if not self.failFunc_list:
    #         print("All functions succeeded. Stopping retries.")
    #         self.retry_timer.stop()
    #         self.ui.pushButton_9.setEnabled(True)
    #         return

    #     self.retry += 1



    # def is_flag_passed(self, func):
    #     if func == self.worker.fun_0x115:
    #         return self.Mains_result == 'Pass'
    #     elif func == self.worker.fun_0x116:
    #         return self.IntVtg_result == 'Pass'
    #     elif func == self.worker.fun_0x109:
    #         return self.Gps_result == 'Pass'
    #     elif func == self.worker.fun_0x110:
    #         return self.GSM_result == 'Pass'
    #     elif func == self.worker.fun_0x121:
    #         return self.Flash_result == 'Pass'
    #     elif func == self.worker.fun_0x114:
    #         return self.MEMS_result == 'Pass'
    #     elif func == self.worker.fun_0x113:
    #         return self.MQTT_result == 'Pass'
    #     elif func == self.worker.fun_0x102:
    #         return self.RTC_result == 'Pass'
    #     elif func == self.worker.fun_0x114:
    #         return self.MEMS_result == 'Pass'
    #     elif func == self.worker.fun_0x123 or self.worker.fun_0x122:
    #         return self.WDT_result == 'Pass'
        
        
        
    #     if self.Mains_result == 'Pass' and self.IntVtg_result == 'Pass' and self.Gps_result == 'Pass' and self.GSM_result == 'Pass' and self.RTC_result == 'Pass' and self.MEMS_result == 'Pass' and self.MQTT_result == 'Pass' and self.Flash_result == 'Pass':
    #             self.ui.textEdit.setPlainText("Pass")
    #             self.ui.textEdit.setStyleSheet("""Font-size:20px; font-weight: Bold; background-color: green""")
                
    #     else:
    #             self.ui.textEdit.setPlainText("Fail")
    #             self.ui.textEdit.setStyleSheet("""Font-size:20px; font-weight: Bold; background-color: red""")
    #     return False

    # def update_ui_on_function_success(self,func):
    #      # Update the UI based on which function passed
    #     if func == self.worker.fun_0x116:
    #         self.ui.IntBat_input_2.setPlainText(self.worker.cleaned_IntBatVtg)
    #         self.ui.plainTextEdit_23.setPlainText("Pass")
    #         self.ui.plainTextEdit_23.setStyleSheet("""Font-size:16px ; font-weight : Bold;background-color:green""")
        
        
    


# # Entry point of the program
# if __name__ == "__main__":
#     app = QApplication(sys.argv)    
#     processor = CAN_Data()
#     processor.show()
#     sys.exit(app.exec_())